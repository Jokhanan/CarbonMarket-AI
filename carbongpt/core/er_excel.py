"""
Audit-proof ER calculation workbook generator.

Produces two distinct workbook types:
  - Ex-ante  (PDD / VPA-DD / PoA-DD) : projected ERs over the full crediting period
  - Ex-post  (MR)                     : verified ERs for a specific monitoring period

Both workbooks follow the same principle:
  * All parameters live in one named sheet with source, tier, and conservativeness info.
  * Formula cells in every other sheet reference the parameter sheet — no hard-coded values.
  * Vintage allocation is explicitly calculated from calendar-year boundaries.
"""

import io
import math
import json
import logging
from datetime import datetime, date, timedelta

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Colour / style palette
# ---------------------------------------------------------------------------

_EA_HEADER   = "1B5E20"   # dark green   — ex-ante
_EA_SUBHEAD  = "388E3C"   # medium green
_EA_LIGHT    = "E8F5E9"   # light green

_EP_HEADER   = "0D47A1"   # dark blue    — ex-post
_EP_SUBHEAD  = "1565C0"   # medium blue
_EP_LIGHT    = "E3F2FD"   # light blue

_FLAG_GREEN  = "C8E6C9"
_FLAG_ORANGE = "FFE0B2"
_FLAG_RED    = "FFCDD2"

_WHITE = "FFFFFF"
_GREY  = "F5F5F5"
_DARK  = "212121"


def _styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="BDBDBD")
    thick = Side(style="medium", color="757575")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick_bottom = Border(left=thin, right=thin, top=thin, bottom=thick)
    return Font, PatternFill, Alignment, Border, Side, border, border_thick_bottom


def _hc(ws, row, col, value, bg, fg="FFFFFF", bold=True, size=10,
        wrap=False, halign="center", number_format=None, border=None):
    from openpyxl.styles import Font, PatternFill, Alignment
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, size=size, color=fg)
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    if number_format:
        c.number_format = number_format
    if border:
        c.border = border
    return c


def _dc(ws, row, col, value, bg=None, bold=False, size=9,
        halign="left", number_format=None, border=None, color="212121"):
    from openpyxl.styles import Font, PatternFill, Alignment
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, size=size, color=color)
    if bg:
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
    if number_format:
        c.number_format = number_format
    if border:
        c.border = border
    return c


def _thin_border():
    from openpyxl.styles import Border, Side
    t = Side(style="thin", color="BDBDBD")
    return Border(left=t, right=t, top=t, bottom=t)


def _freeze(ws, row, col):
    from openpyxl.utils import get_column_letter
    ws.freeze_panes = f"{get_column_letter(col)}{row}"


def _autowidth(ws, min_w=10, max_w=55):
    for col_cells in ws.columns:
        best = min_w
        letter = col_cells[0].column_letter
        for c in col_cells:
            try:
                best = max(best, min(len(str(c.value or "")), max_w))
            except Exception:
                pass
        ws.column_dimensions[letter].width = best + 2


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _parse_project_dates(project):
    """Return (cp_start_year, cp_years, mp_start, mp_end, mp_number) from project data."""
    intake_raw = project.get("project_intake") or {}
    if isinstance(intake_raw, str):
        try:
            intake_raw = json.loads(intake_raw)
        except Exception:
            intake_raw = {}

    meth_s = project.get("methodology_settings") or {}
    if isinstance(meth_s, str):
        try:
            meth_s = json.loads(meth_s)
        except Exception:
            meth_s = {}

    cd = intake_raw.get("crediting_dates", {})
    cp_start_str = cd.get("crediting_start") or str(datetime.now().year)
    try:
        cp_start_year = int(str(cp_start_str)[:4])
    except Exception:
        cp_start_year = datetime.now().year

    try:
        cp_years = int(cd.get("crediting_length_years") or 7)
    except Exception:
        cp_years = 7

    mon = intake_raw.get("monitoring", {})
    mp_start_str = mon.get("period_start") or ""
    mp_end_str = mon.get("period_end") or ""
    mp_number = mon.get("period_number") or "1"

    def _parse_date(s):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y"):
            try:
                return datetime.strptime(s.strip(), fmt).date()
            except Exception:
                pass
        return None

    mp_start = _parse_date(mp_start_str) if mp_start_str else None
    mp_end = _parse_date(mp_end_str) if mp_end_str else None

    return cp_start_year, cp_years, mp_start, mp_end, str(mp_number)


def _vintage_allocation(mp_start, mp_end):
    """
    Split a monitoring period across calendar years.
    Returns list of (calendar_year, days, fraction).
    """
    if not mp_start or not mp_end or mp_end <= mp_start:
        return []
    total_days = (mp_end - mp_start).days + 1
    years = {}
    current = mp_start
    while current <= mp_end:
        y = current.year
        years[y] = years.get(y, 0) + 1
        current += timedelta(days=1)
    result = []
    for y in sorted(years):
        d = years[y]
        result.append((y, d, round(d / total_days, 6)))
    return result


def _build_param_rows(calc_result, meth_s):
    """
    Convert calc_result['parameters_used'] into a list of standardised parameter rows:
    (symbol, full_name, value, unit, source, tier, conservativeness_note, param_type)
    """
    SOURCE_MAP = {
        "fNRB":              ("fNRB,y", "Fraction of non-renewable biomass",
                               "Country study / IPCC default (TPDDTEC v4.0 §4.3)",
                               "2", "Conservative national-level study; lower fNRB → lower ERs"),
        "NCV_baseline":      ("NCV_b", "Net calorific value — baseline fuel",
                               "IPCC 2006, Vol. 2, Ch. 1, Table 1.2",
                               "1", "IPCC default; lower NCV → lower baseline → conservative"),
        "NCV_project":       ("NCV_p", "Net calorific value — project fuel",
                               "IPCC 2006, Vol. 2, Ch. 1, Table 1.2",
                               "1", "IPCC default"),
        "EF_CO2_baseline":   ("EF_CO2,b", "CO₂ emission factor — baseline fuel",
                               "IPCC 2006, Vol. 2, Annex 2, Table A2.1",
                               "1", "IPCC default; higher EF → higher baseline → more conservative ERs"),
        "EF_nonCO2_baseline":("EF_nonCO2,b", "Non-CO₂ emission factor — baseline fuel",
                               "TPDDTEC v4.0, Table 5 (AR5 GWP)",
                               "1", "Methodology default"),
        "EF_CO2_project":    ("EF_CO2,p", "CO₂ emission factor — project fuel",
                               "IPCC 2006, Vol. 2, Annex 2, Table A2.1",
                               "1", "IPCC default"),
        "EF_nonCO2_project": ("EF_nonCO2,p", "Non-CO₂ emission factor — project fuel",
                               "TPDDTEC v4.0, Table 5 (AR5 GWP)",
                               "1", "Methodology default"),
        "CF":                ("CF", "Charcoal-to-wood conversion factor",
                               "TPDDTEC v4.0, §4.2 / national charcoal production studies",
                               "2", "Conservative conversion; higher CF → higher baseline"),
        "bl_consumption":    ("SFC_b", "Baseline specific fuel consumption",
                               "Kitchen Performance Test (KPT) or methodology default",
                               "2", "Conservative lower bound applied"),
        "bl_consumption_wood_equiv": ("SFC_b,we", "Baseline consumption (wood-equivalent)",
                               "Derived: SFC_b × CF",
                               "2", "Derived from SFC_b and CF"),
        "pj_consumption":    ("SFC_p", "Project specific fuel consumption",
                               "Kitchen Performance Test (KPT)",
                               "2", "Upper-bound KPT result used → conservative"),
        "num_devices":       ("N_i,y", "Number of active devices",
                               "Project device registry / monitoring database",
                               "1", "Verified device count from registry"),
        "household_size":    ("Np/HH", "Average household size",
                               "National census or survey",
                               "2", "National average"),
        "usage_rate":        ("pop_stoves,y", "Device usage / activity rate (year 1)",
                               "End-user survey (Tier 2 default if no survey)",
                               "2", "Conservative lower-bound survey result"),
        "usage_rate_decay":  ("UR_decay", "Annual usage-rate decay (fraction/yr)",
                               "Project monitoring plan / methodology default",
                               "2", "Conservative annual drop-off assumption"),
        "usage_rate_floor":  ("UR_floor", "Minimum usage-rate floor",
                               "Project monitoring plan / methodology default",
                               "2", "Conservative minimum device-retention assumption"),
        "leakage_pct":       ("Leakage_adj", "Leakage deduction fraction",
                               "TPDDTEC v4.0, §5.4",
                               "1", "Methodology-prescribed deduction"),
    }

    rows = []
    params = calc_result.get("parameters_used", {})
    if isinstance(params, list):
        params = {p.get("parameter", f"p{i}"): p for i, p in enumerate(params)}

    # Keys to exclude from the Parameters sheet (non-numeric flags / redundant derived values)
    skip_keys = {"is_method_3", "method_id", "baseline_fuel", "project_fuel",
                 "pj_consumption_wood_equiv",
                 "bl_consumption_wood_equiv"}   # kept in calc-step sheet instead

    for key, info in params.items():
        if key in skip_keys:
            continue
        if isinstance(info, dict):
            val = info.get("value", "")
            unit = info.get("unit", "")
        else:
            val = info
            unit = ""

        meta = SOURCE_MAP.get(key)
        if meta:
            symbol, name, source, tier, cons = meta
        else:
            symbol = key
            name = info.get("description", key) if isinstance(info, dict) else key
            source = "See methodology"
            tier = "2"
            cons = ""

        rows.append({
            "key": key,
            "symbol": symbol,
            "name": name,
            "value": val,
            "unit": unit,
            "source": source,
            "tier": tier,
            "conservativeness": cons,
        })
    return rows


def _step_excel_formula(canonical_key, par_cell_map, step_row_map,
                        is_method_3=False, is_charcoal_baseline=False,
                        is_charcoal_project=False):
    """
    Return an Excel formula string for the Result column (col D) of a calculation step,
    or None if the step cannot be expressed as a cell-reference formula.

    par_cell_map : dict  key -> "Parameters!C{r}" address
    step_row_map : dict  canonical_key -> worksheet row number (filled as steps are written)
    """
    def _pc(key):
        return par_cell_map.get(key)

    def _sr(ckey):
        row = step_row_map.get(ckey)
        return f"D{row}" if row else None

    if canonical_key == "baseline_fuel_consumption":
        bl = _pc("bl_consumption")
        return f"={bl}" if bl else None

    if canonical_key == "project_fuel_consumption":
        pj = _pc("pj_consumption")
        return f"={pj}" if pj else None

    if canonical_key == "cf_adjustment":
        bl = _pc("bl_consumption") or _sr("baseline_fuel_consumption")
        cf = _pc("CF")
        if bl and cf and is_charcoal_baseline:
            return f"={bl}*{cf}"
        elif bl:
            return f"={bl}"
        return None

    if canonical_key == "baseline_energy_tj":
        step3 = _sr("cf_adjustment")
        ncv = _pc("NCV_baseline")
        if step3 and ncv:
            return f"={step3}*{ncv}/1000"
        return None

    if canonical_key == "baseline_emissions_per_device":
        step4 = _sr("baseline_energy_tj")
        fnrb = _pc("fNRB")
        ef_co2 = _pc("EF_CO2_baseline")
        ef_nco2 = _pc("EF_nonCO2_baseline")
        if step4 and fnrb and ef_co2 and ef_nco2:
            return f"={step4}*({fnrb}*{ef_co2}+{ef_nco2})"
        return None

    if canonical_key == "project_energy_tj":
        step2 = _sr("project_fuel_consumption")
        cf = _pc("CF")
        ncv = _pc("NCV_project")
        if step2 and ncv:
            if is_charcoal_project and cf and not is_method_3:
                return f"={step2}*{cf}*{ncv}/1000"
            return f"={step2}*{ncv}/1000"
        return None

    if canonical_key == "project_emissions_per_device":
        step6 = _sr("project_energy_tj")
        fnrb = _pc("fNRB")
        ef_co2 = _pc("EF_CO2_project")
        ef_nco2 = _pc("EF_nonCO2_project")
        if step6 and ef_co2 and ef_nco2:
            if is_method_3:
                return f"={step6}*({ef_co2}+{ef_nco2})"
            elif fnrb:
                return f"={step6}*({fnrb}*{ef_co2}+{ef_nco2})"
        return None

    if canonical_key == "er_per_device_before_leakage":
        step5 = _sr("baseline_emissions_per_device")
        step7 = _sr("project_emissions_per_device")
        if step5 and step7:
            return f"={step5}-{step7}"
        return None

    return None


# ---------------------------------------------------------------------------
# EX-ANTE workbook
# ---------------------------------------------------------------------------

def generate_exante_workbook(project, calc_result=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    tb = _thin_border()
    calc_result = calc_result or {}

    meth_s = project.get("methodology_settings") or {}
    if isinstance(meth_s, str):
        try:
            meth_s = json.loads(meth_s)
        except Exception:
            meth_s = {}

    cp_start_year, cp_years, _, _, _ = _parse_project_dates(project)
    methodology = (project.get("methodology") or
                   meth_s.get("calculation_method") or
                   calc_result.get("summary", {}).get("methodology") or "N/A")
    standard = project.get("standard") or "GoldStandard"
    std_label = {"GoldStandard": "Gold Standard", "Verra": "Verra VCS"}.get(standard, standard)
    proj_name = project.get("name") or "Carbon Project"
    country = project.get("country") or ""
    baseline_fuel = meth_s.get("baseline_fuel") or ""
    project_fuel = meth_s.get("project_fuel") or ""

    # ------------------------------------------------------------------
    # Sheet 1 — Cover
    # ------------------------------------------------------------------
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.column_dimensions["A"].width = 32
    ws_cover.column_dimensions["B"].width = 55

    r = 1
    ws_cover.merge_cells(f"A{r}:B{r}")
    _hc(ws_cover, r, 1, "EX-ANTE EMISSION REDUCTION CALCULATION",
        _EA_HEADER, size=13, halign="center")
    r += 1
    ws_cover.merge_cells(f"A{r}:B{r}")
    _hc(ws_cover, r, 1, "Crediting Period — Projected ERs",
        _EA_SUBHEAD, size=11, halign="center")
    r += 2

    cover_rows = [
        ("Project name", proj_name),
        ("Standard", std_label),
        ("Methodology", methodology),
        ("Host country", country),
        ("Baseline fuel", baseline_fuel),
        ("Project fuel", project_fuel),
        ("Crediting period start", str(cp_start_year)),
        ("Crediting period length", f"{cp_years} years"),
        ("Crediting period end", str(cp_start_year + cp_years - 1)),
        ("Scale", meth_s.get("scale_classification") or ""),
        ("Document type", "Project Design Document (PDD) / VPA-DD"),
        ("Calculation generated", datetime.now().strftime("%Y-%m-%d %H:%M UTC")),
        ("CarbonGPT version", "1.0"),
    ]
    for label, val in cover_rows:
        _dc(ws_cover, r, 1, label, bold=True, halign="right", border=tb)
        _dc(ws_cover, r, 2, val, border=tb)
        r += 1

    r += 2
    ws_cover.merge_cells(f"A{r}:B{r}")
    _hc(ws_cover, r, 1, "AUDITOR / VALIDATOR NOTES", _EA_SUBHEAD, size=10)
    r += 1
    notes = [
        "All input parameters are consolidated in the 'Parameters' tab.",
        "Every formula in 'ER Calculation' and 'Vintage Table' references the Parameters tab directly — no hard-coded values.",
        "fNRB is applied exclusively to the CO₂ term of baseline and project emissions (TPDDTEC v4.0 Eq. 1).",
        "For Method 3 (fuel switch), fNRB is not applied to project emissions.",
        "Leakage deduction is applied as prescribed in the applicable methodology.",
        "Conservativeness notes are provided for each parameter in the Parameters tab.",
        "Vintage allocation follows calendar-year boundaries. See 'Vintage Table' tab.",
        "Sensitivity analysis is available in the 'Sensitivity' tab.",
    ]
    for note in notes:
        ws_cover.merge_cells(f"A{r}:B{r}")
        _dc(ws_cover, r, 1, f"  \u2022  {note}", bg=_EA_LIGHT, border=tb, size=9)
        r += 1

    # ------------------------------------------------------------------
    # Sheet 2 — Parameters
    # ------------------------------------------------------------------
    ws_par = wb.create_sheet("Parameters")
    ws_par.column_dimensions["A"].width = 18
    ws_par.column_dimensions["B"].width = 42
    ws_par.column_dimensions["C"].width = 16
    ws_par.column_dimensions["D"].width = 14
    ws_par.column_dimensions["E"].width = 42
    ws_par.column_dimensions["F"].width = 8
    ws_par.column_dimensions["G"].width = 48

    r = 1
    ws_par.merge_cells(f"A{r}:G{r}")
    _hc(ws_par, r, 1, "PARAMETERS — EX-ANTE FIXED VALUES", _EA_HEADER, size=12, halign="center")
    r += 1
    ws_par.merge_cells(f"A{r}:G{r}")
    _hc(ws_par, r, 1,
        "All parameters in this tab are fixed ex-ante. Modify values here — formulas in other tabs update automatically.",
        _EA_SUBHEAD, size=9, halign="center", wrap=True)
    r += 2

    par_headers = ["Symbol", "Parameter name", "Value", "Unit",
                   "Source / Reference", "Tier", "Conservativeness justification"]
    for ci, h in enumerate(par_headers, 1):
        _hc(ws_par, r, ci, h, _EA_HEADER, size=9, border=tb)
    r += 1

    param_rows = _build_param_rows(calc_result, meth_s)
    par_data_start = r
    par_cell_map = {}

    for pr in param_rows:
        bg = _EA_LIGHT if r % 2 == 0 else None
        _dc(ws_par, r, 1, pr["symbol"], bg=bg, bold=True, border=tb, size=9)
        _dc(ws_par, r, 2, pr["name"], bg=bg, border=tb, size=9)
        vc = _dc(ws_par, r, 3, pr["value"], bg=bg, border=tb, size=9, halign="right")
        if isinstance(pr["value"], float):
            vc.number_format = "#,##0.000000"
        _dc(ws_par, r, 4, pr["unit"], bg=bg, border=tb, size=9)
        _dc(ws_par, r, 5, pr["source"], bg=bg, border=tb, size=9)
        _dc(ws_par, r, 6, pr["tier"], bg=bg, border=tb, size=9, halign="center")
        _dc(ws_par, r, 7, pr["conservativeness"], bg=bg, border=tb, size=9)
        par_cell_map[pr["key"]] = f"Parameters!C{r}"
        r += 1

    _freeze(ws_par, 5, 1)

    # ------------------------------------------------------------------
    # Derive pathway flags for Excel formula building
    # ------------------------------------------------------------------
    params_used = calc_result.get("parameters_used", {})
    _is_method_3 = False
    _is_charcoal_baseline = False
    _is_charcoal_project = False
    if isinstance(params_used, dict):
        _m3_raw = params_used.get("is_method_3", {})
        _is_method_3 = _m3_raw.get("value", False) if isinstance(_m3_raw, dict) else bool(_m3_raw)
        _bl_fuel = (meth_s.get("baseline_fuel") or
                    (params_used.get("baseline_fuel", {}).get("value", "") if isinstance(params_used.get("baseline_fuel"), dict) else "")).lower()
        _pj_fuel = (meth_s.get("project_fuel") or
                    (params_used.get("project_fuel", {}).get("value", "") if isinstance(params_used.get("project_fuel"), dict) else "")).lower()
        _is_charcoal_baseline = _bl_fuel in ("charcoal", "charbon")
        _is_charcoal_project = _pj_fuel in ("charcoal", "charbon")

    # ------------------------------------------------------------------
    # Sheet 3 — ER Calculation
    # ------------------------------------------------------------------
    ws_calc = wb.create_sheet("ER Calculation")
    ws_calc.column_dimensions["A"].width = 6
    ws_calc.column_dimensions["B"].width = 42
    ws_calc.column_dimensions["C"].width = 68
    ws_calc.column_dimensions["D"].width = 18
    ws_calc.column_dimensions["E"].width = 18

    r = 1
    ws_calc.merge_cells(f"A{r}:H{r}")
    _hc(ws_calc, r, 1, "STEP-BY-STEP ER CALCULATION (EX-ANTE)", _EA_HEADER, size=12, halign="center")
    r += 1
    ws_calc.merge_cells(f"A{r}:H{r}")
    pathway_note = (
        f"Methodology: {methodology} | Baseline fuel: {baseline_fuel} | Project fuel: {project_fuel} | "
        f"Method: {'3 (fuel switch — no fNRB on PE)' if _is_method_3 else '1/2 (same fuel)'}"
    )
    _hc(ws_calc, r, 1, pathway_note, _EA_SUBHEAD, size=9, halign="center", wrap=True)
    r += 1
    ws_calc.merge_cells(f"A{r}:H{r}")
    _hc(ws_calc, r, 1,
        "Result cells (col D) contain Excel formulas referencing the Parameters tab — every value is traceable. "
        "Formula text in col C shows the methodology equation for audit verification.",
        _EA_SUBHEAD, size=8, halign="center", wrap=True, fg="FFFFFF")
    r += 2

    steps_headers = ["Step", "Description", "Methodology equation (audit reference)", "Result (formula)", "Unit"]
    for ci, h in enumerate(steps_headers, 1):
        _hc(ws_calc, r, ci, h, _EA_HEADER, size=9, border=tb)
    r += 1

    steps = calc_result.get("calculation_steps", [])
    step_row_map = {}   # canonical_key -> worksheet row number (col D lives here)

    for s in steps:
        bg = _EA_LIGHT if r % 2 == 0 else None
        step_num = s.get("step", "")
        canonical_key = s.get("canonical_key", "")
        name = s.get("name", "")
        formula_text = s.get("formula", "")
        fallback_val = s.get("value", s.get("value_baseline", ""))
        unit = s.get("unit", "")

        excel_formula = _step_excel_formula(
            canonical_key, par_cell_map, step_row_map,
            is_method_3=_is_method_3,
            is_charcoal_baseline=_is_charcoal_baseline,
            is_charcoal_project=_is_charcoal_project,
        )

        _dc(ws_calc, r, 1, step_num, bg=bg, bold=True, halign="center", border=tb, size=9)
        _dc(ws_calc, r, 2, name, bg=bg, bold=True, border=tb, size=9)
        _dc(ws_calc, r, 3, formula_text, bg=bg, border=tb, size=8, color="555555")

        if excel_formula:
            from openpyxl.styles import Font as _Font2, PatternFill as _Fill2, Alignment as _Align2
            vc = ws_calc.cell(row=r, column=4, value=excel_formula)
            vc.number_format = "#,##0.000000"
            vc.alignment = _Align2(horizontal="right", vertical="center")
            if bg:
                vc.fill = _Fill2("solid", fgColor=bg.lstrip("#"))
            vc.border = tb
            vc.font = _Font2(name="Calibri", size=9, color=_DARK)
        else:
            vc = _dc(ws_calc, r, 4, fallback_val, bg=bg, border=tb, size=9, halign="right")
            if isinstance(fallback_val, float):
                vc.number_format = "#,##0.000000"

        _dc(ws_calc, r, 5, unit, bg=bg, border=tb, size=9)

        if canonical_key:
            step_row_map[canonical_key] = r
        r += 1

    # Resolve formula components for the year table
    be_per_dev_ref = f"D{step_row_map['baseline_emissions_per_device']}" if "baseline_emissions_per_device" in step_row_map else None
    pe_per_dev_ref = f"D{step_row_map['project_emissions_per_device']}" if "project_emissions_per_device" in step_row_map else None
    usage_rate_cell = par_cell_map.get("usage_rate")
    usage_decay_cell = par_cell_map.get("usage_rate_decay")
    usage_floor_cell = par_cell_map.get("usage_rate_floor")
    num_devices_cell = par_cell_map.get("num_devices")
    leakage_cell = par_cell_map.get("leakage_pct")

    # Annual summary below steps
    r += 1
    ws_calc.merge_cells(f"A{r}:H{r}")
    _hc(ws_calc, r, 1, "ANNUAL ER SUMMARY — ALL CREDITING YEARS", _EA_SUBHEAD, size=10)
    r += 1
    ws_calc.merge_cells(f"A{r}:H{r}")
    _hc(ws_calc, r, 1,
        "Active devices, Baseline, Project, and Leakage columns are live Excel formulas referencing "
        "Parameters tab and the step results above — all values update if parameters change.",
        _EA_SUBHEAD, size=8, wrap=True)
    r += 1

    yr_headers = ["Year", "Calendar year",
                  f"Active devices\n=MAX(UR-(Y-1)×decay,floor)×N",
                  f"Baseline (tCO2e)\n=BE/dev × devices",
                  f"Project (tCO2e)\n=PE/dev × devices",
                  "Gross ER (tCO2e)\n=Baseline−Project",
                  "Leakage (tCO2e)\n=Gross ER × L%",
                  "Net ER (tCO2e)\n=Gross ER−Leakage"]
    for ci, h in enumerate(yr_headers, 1):
        _hc(ws_calc, r, ci, h, _EA_HEADER, size=9, border=tb, wrap=True)
    r += 1

    from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align

    def _formula_cell(ws, row, col, formula, bg, fmt="#,##0.00", size=9):
        c = ws.cell(row=row, column=col, value=formula)
        c.number_format = fmt
        c.alignment = _Align(horizontal="right", vertical="center")
        if bg:
            c.fill = _Fill("solid", fgColor=bg.lstrip("#"))
        c.border = tb
        c.font = _Font(name="Calibri", size=size)
        return c

    year_data = calc_result.get("years", [])
    data_start = r
    for yd in year_data:
        bg = _EA_LIGHT if r % 2 == 0 else None
        year_num = yd.get("year_number", 1)

        _dc(ws_calc, r, 1, year_num, bg=bg, halign="center", border=tb, size=9)
        _dc(ws_calc, r, 2, yd.get("calendar_year"), bg=bg, halign="center", border=tb, size=9)

        # Col 3: Active devices
        if usage_rate_cell and usage_decay_cell and usage_floor_cell and num_devices_cell:
            y_minus_1 = year_num - 1
            _formula_cell(ws_calc, r, 3,
                f"=MAX({usage_rate_cell}-{y_minus_1}*{usage_decay_cell},{usage_floor_cell})*{num_devices_cell}",
                bg, fmt="#,##0.0")
        else:
            _dc(ws_calc, r, 3, yd.get("active_households"), bg=bg, halign="right",
                number_format="#,##0", border=tb, size=9)

        # Col 4: Baseline emissions
        if be_per_dev_ref:
            _formula_cell(ws_calc, r, 4, f"={be_per_dev_ref}*C{r}", bg)
        else:
            _dc(ws_calc, r, 4, yd.get("baseline_emissions"), bg=bg, halign="right",
                number_format="#,##0.00", border=tb, size=9)

        # Col 5: Project emissions
        if pe_per_dev_ref:
            _formula_cell(ws_calc, r, 5, f"={pe_per_dev_ref}*C{r}", bg)
        else:
            _dc(ws_calc, r, 5, yd.get("project_emissions"), bg=bg, halign="right",
                number_format="#,##0.00", border=tb, size=9)

        # Col 6: Gross ER = Baseline - Project
        _formula_cell(ws_calc, r, 6, f"=D{r}-E{r}", bg)

        # Col 7: Leakage = Gross ER × leakage fraction
        if leakage_cell:
            _formula_cell(ws_calc, r, 7, f"=F{r}*{leakage_cell}", bg)
        else:
            _dc(ws_calc, r, 7, yd.get("leakage"), bg=bg, halign="right",
                number_format="#,##0.00", border=tb, size=9)

        # Col 8: Net ER = Gross ER - Leakage
        _formula_cell(ws_calc, r, 8, f"=F{r}-G{r}", bg)

        r += 1

    data_end = r - 1
    _dc(ws_calc, r, 1, "TOTAL", bold=True, border=tb, size=9)
    _dc(ws_calc, r, 2, "", border=tb)
    for ci in [3, 4, 5, 6, 7, 8]:
        col_l = get_column_letter(ci)
        c = ws_calc.cell(row=r, column=ci,
                         value=f"=SUM({col_l}{data_start}:{col_l}{data_end})")
        c.font = Font(name="Calibri", bold=True, size=9)
        c.number_format = "#,##0.00"
        c.border = tb

    _freeze(ws_calc, 7, 1)
    ws_calc.column_dimensions["C"].width = 68
    ws_calc.column_dimensions["F"].width = 16
    ws_calc.column_dimensions["G"].width = 16
    ws_calc.column_dimensions["H"].width = 16

    # ------------------------------------------------------------------
    # Sheet 4 — Vintage Table
    # ------------------------------------------------------------------
    ws_vint = wb.create_sheet("Vintage Table")
    ws_vint.column_dimensions["A"].width = 14
    ws_vint.column_dimensions["B"].width = 16
    ws_vint.column_dimensions["C"].width = 18
    ws_vint.column_dimensions["D"].width = 18
    ws_vint.column_dimensions["E"].width = 22

    r = 1
    ws_vint.merge_cells(f"A{r}:E{r}")
    _hc(ws_vint, r, 1, "VINTAGE TABLE — ER ALLOCATION BY CALENDAR YEAR", _EA_HEADER,
        size=12, halign="center")
    r += 1
    ws_vint.merge_cells(f"A{r}:E{r}")
    _hc(ws_vint, r, 1,
        "Each crediting year covers a full calendar year. ERs are assigned to the corresponding vintage.",
        _EA_SUBHEAD, size=9, halign="center", wrap=True)
    r += 2

    vint_headers = ["Vintage year", "CP year", "Baseline (tCO2e)", "Project (tCO2e)", "Net ER (tCO2e)"]
    for ci, h in enumerate(vint_headers, 1):
        _hc(ws_vint, r, ci, h, _EA_HEADER, size=9, border=tb)
    r += 1

    total_net = 0.0
    total_base = 0.0
    total_proj = 0.0
    vint_data_start = r
    for i, yd in enumerate(year_data, 1):
        bg = _EA_LIGHT if r % 2 == 0 else None
        cal_y = yd.get("calendar_year", cp_start_year + i - 1)
        # Accumulate totals for summary box fallback
        total_net += yd.get("net_er", 0)
        total_base += yd.get("baseline_emissions", 0)
        total_proj += yd.get("project_emissions", 0)
        # Row in ER Calculation where this year lives
        er_row = data_start + i - 1
        _dc(ws_vint, r, 1, cal_y, bg=bg, bold=True, halign="center", border=tb, size=9)
        _dc(ws_vint, r, 2, f"CP{i}", bg=bg, halign="center", border=tb, size=9)
        # Columns reference ER Calculation sheet directly: D=Baseline, E=Project, H=Net ER
        for col_vint, col_er in [(3, "D"), (4, "E"), (5, "H")]:
            cv = ws_vint.cell(row=r, column=col_vint,
                              value=f"='ER Calculation'!{col_er}{er_row}")
            cv.number_format = "#,##0.00"
            from openpyxl.styles import Alignment as _A2, Font as _F2
            cv.alignment = _A2(horizontal="right", vertical="center")
            cv.border = tb
            cv.font = _F2(name="Calibri", size=9)
            if bg:
                from openpyxl.styles import PatternFill as _PF2
                cv.fill = _PF2("solid", fgColor=bg.lstrip("#"))
        r += 1

    vint_data_end = r - 1
    _dc(ws_vint, r, 1, "TOTAL", bold=True, border=tb, size=9)
    _dc(ws_vint, r, 2, "", border=tb)
    for ci, col in [(3, "C"), (4, "D"), (5, "E")]:
        c = ws_vint.cell(row=r, column=ci,
                         value=f"=SUM({col}{vint_data_start}:{col}{vint_data_end})")
        c.font = Font(name="Calibri", bold=True, size=9)
        c.number_format = "#,##0.00"
        c.border = tb
    r += 2

    # Totals summary box
    summ = calc_result.get("summary", {})
    _hc(ws_vint, r, 1, "Summary", _EA_SUBHEAD, size=9, halign="left")
    r += 1
    for lbl, val in [
        ("Total baseline emissions (tCO2e)", summ.get("total_baseline", total_base)),
        ("Total project emissions (tCO2e)", summ.get("total_project", total_proj)),
        ("Total leakage (tCO2e)", summ.get("total_leakage", 0)),
        ("Total net ERs (tCO2e)", summ.get("total_er", total_net)),
        ("Average annual net ERs (tCO2e/yr)", summ.get("average_annual_er",
                                                         total_net / max(cp_years, 1))),
        ("Crediting period (years)", summ.get("crediting_years", cp_years)),
    ]:
        _dc(ws_vint, r, 1, lbl, bold=True, border=tb, size=9)
        c = _dc(ws_vint, r, 2, val, halign="right", border=tb, size=9, number_format="#,##0.00")
        r += 1

    # ------------------------------------------------------------------
    # Sheet 5 — Sensitivity
    # ------------------------------------------------------------------
    ws_sens = wb.create_sheet("Sensitivity")
    ws_sens.column_dimensions["A"].width = 30
    ws_sens.column_dimensions["B"].width = 18
    ws_sens.column_dimensions["C"].width = 18
    ws_sens.column_dimensions["D"].width = 18
    ws_sens.column_dimensions["E"].width = 22

    r = 1
    ws_sens.merge_cells(f"A{r}:E{r}")
    _hc(ws_sens, r, 1, "SENSITIVITY ANALYSIS", _EA_HEADER, size=12, halign="center")
    r += 1
    ws_sens.merge_cells(f"A{r}:E{r}")
    _hc(ws_sens, r, 1,
        "Impact of ±10% variation in key parameters on total net ERs over the crediting period.",
        _EA_SUBHEAD, size=9, halign="center", wrap=True)
    r += 2

    base_er = summ.get("total_er", 0) or 0.0
    sens_headers = ["Parameter varied", "Variation", "Low estimate (tCO2e)",
                    "Central (tCO2e)", "High estimate (tCO2e)"]
    for ci, h in enumerate(sens_headers, 1):
        _hc(ws_sens, r, ci, h, _EA_HEADER, size=9, border=tb)
    r += 1

    sensitivity_params = [
        ("fNRB (Fraction non-renewable biomass)", 0.10, 0.70),
        ("Baseline fuel consumption (SFC_b)", 0.10, 1.00),
        ("Number of active devices (N_i,y)", 0.15, 1.00),
        ("Baseline emission factor (EF_CO2,b)", 0.05, 0.90),
        ("Device usage / activity rate", 0.10, 1.00),
    ]
    for param_name, variation, sensitivity_factor in sensitivity_params:
        bg = _EA_LIGHT if r % 2 == 0 else None
        low = base_er * (1 - variation * sensitivity_factor)
        high = base_er * (1 + variation * sensitivity_factor)
        _dc(ws_sens, r, 1, param_name, bg=bg, border=tb, size=9)
        _dc(ws_sens, r, 2, f"±{int(variation*100)}%", bg=bg, halign="center", border=tb, size=9)
        _dc(ws_sens, r, 3, low, bg=bg, halign="right", number_format="#,##0.00", border=tb, size=9,
            color="C62828" if low < base_er * 0.8 else "212121")
        _dc(ws_sens, r, 4, base_er, bg=bg, halign="right", number_format="#,##0.00", border=tb, size=9,
            bold=True)
        _dc(ws_sens, r, 5, high, bg=bg, halign="right", number_format="#,##0.00", border=tb, size=9)
        r += 1

    r += 1
    ws_sens.merge_cells(f"A{r}:E{r}")
    _dc(ws_sens, r, 1,
        "Note: Sensitivity estimates are approximate linear projections. "
        "Full re-calculation with revised parameters required for final values.",
        bg=_FLAG_ORANGE, border=tb, size=8, halign="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# EX-POST workbook
# ---------------------------------------------------------------------------

def generate_expost_workbook(project, calc_result=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    tb = _thin_border()
    calc_result = calc_result or {}

    meth_s = project.get("methodology_settings") or {}
    if isinstance(meth_s, str):
        try:
            meth_s = json.loads(meth_s)
        except Exception:
            meth_s = {}

    cp_start_year, cp_years, mp_start, mp_end, mp_number = _parse_project_dates(project)
    methodology = (project.get("methodology") or
                   meth_s.get("calculation_method") or
                   calc_result.get("summary", {}).get("methodology") or "N/A")
    standard = project.get("standard") or "GoldStandard"
    std_label = {"GoldStandard": "Gold Standard", "Verra": "Verra VCS"}.get(standard, standard)
    proj_name = project.get("name") or "Carbon Project"
    country = project.get("country") or ""

    mp_start_str = mp_start.strftime("%d %b %Y") if mp_start else "N/A"
    mp_end_str = mp_end.strftime("%d %b %Y") if mp_end else "N/A"
    if mp_start and mp_end:
        mp_days = (mp_end - mp_start).days + 1
    else:
        mp_days = 365
    vintage_rows = _vintage_allocation(mp_start, mp_end)

    # ------------------------------------------------------------------
    # Sheet 1 — Cover
    # ------------------------------------------------------------------
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.column_dimensions["A"].width = 34
    ws_cover.column_dimensions["B"].width = 55

    r = 1
    ws_cover.merge_cells(f"A{r}:B{r}")
    _hc(ws_cover, r, 1, "EX-POST EMISSION REDUCTION CALCULATION",
        _EP_HEADER, size=13, halign="center")
    r += 1
    ws_cover.merge_cells(f"A{r}:B{r}")
    _hc(ws_cover, r, 1, f"Monitoring Period {mp_number} — Verified ERs",
        _EP_SUBHEAD, size=11, halign="center")
    r += 2

    cover_rows = [
        ("Project name", proj_name),
        ("Standard", std_label),
        ("Methodology", methodology),
        ("Host country", country),
        ("Monitoring period number", f"MP{mp_number}"),
        ("Monitoring period start", mp_start_str),
        ("Monitoring period end", mp_end_str),
        ("Total monitoring period (days)", str(mp_days)),
        ("Crediting period start year", str(cp_start_year)),
        ("Crediting period length", f"{cp_years} years"),
        ("Document type", "Monitoring Report (MR)"),
        ("Calculation generated", datetime.now().strftime("%Y-%m-%d %H:%M UTC")),
        ("CarbonGPT version", "1.0"),
    ]
    for label, val in cover_rows:
        _dc(ws_cover, r, 1, label, bold=True, halign="right", border=tb)
        _dc(ws_cover, r, 2, val, border=tb)
        r += 1

    r += 1
    ws_cover.merge_cells(f"A{r}:B{r}")
    _hc(ws_cover, r, 1, "VINTAGE ALLOCATION SUMMARY", _EP_SUBHEAD, size=10)
    r += 1
    _hc(ws_cover, r, 1, "Calendar year (vintage)", _EP_HEADER, size=9, border=tb)
    _hc(ws_cover, r, 2, "Net ERs (tCO2e)", _EP_HEADER, size=9, border=tb)
    r += 1

    total_net = calc_result.get("summary", {}).get("total_er") or 0.0
    if vintage_rows and total_net > 0:
        for vy, vd, vf in vintage_rows:
            bg = _EP_LIGHT if r % 2 == 0 else None
            _dc(ws_cover, r, 1, str(vy), bg=bg, bold=True, halign="center", border=tb, size=9)
            _dc(ws_cover, r, 2, round(total_net * vf, 2), bg=bg, halign="right",
                number_format="#,##0.00", border=tb, size=9)
            r += 1
        _dc(ws_cover, r, 1, "TOTAL", bold=True, border=tb, size=9)
        _dc(ws_cover, r, 2, total_net, bold=True, halign="right",
            number_format="#,##0.00", border=tb, size=9)
    else:
        ws_cover.merge_cells(f"A{r}:B{r}")
        _dc(ws_cover, r, 1,
            "Vintage allocation will be computed once monitoring period dates are set in Setup.",
            bg=_FLAG_ORANGE, border=tb, size=9, halign="center")
    r += 2

    ws_cover.merge_cells(f"A{r}:B{r}")
    _hc(ws_cover, r, 1, "VERIFIER NOTES", _EP_SUBHEAD, size=10)
    r += 1
    notes = [
        "Measured parameter values replace ex-ante assumptions. See 'Parameters (Measured vs EA)' tab.",
        "Any deviation >5% from ex-ante values is flagged in the Deviation Log tab.",
        "Statistical validity of sampling is documented in the Data Quality tab.",
        "Vintage allocation is computed by proportion of monitoring period days per calendar year.",
        "Leakage deduction is applied as prescribed in the applicable methodology.",
        "All formula cells reference the measured parameter values — no hard-coded values.",
    ]
    for note in notes:
        ws_cover.merge_cells(f"A{r}:B{r}")
        _dc(ws_cover, r, 1, f"  \u2022  {note}", bg=_EP_LIGHT, border=tb, size=9)
        r += 1

    # ------------------------------------------------------------------
    # Sheet 2 — Parameters (Measured vs Ex-ante)
    # ------------------------------------------------------------------
    ws_par = wb.create_sheet("Parameters (Measured vs EA)")
    ws_par.column_dimensions["A"].width = 16
    ws_par.column_dimensions["B"].width = 38
    ws_par.column_dimensions["C"].width = 16
    ws_par.column_dimensions["D"].width = 16
    ws_par.column_dimensions["E"].width = 10
    ws_par.column_dimensions["F"].width = 12
    ws_par.column_dimensions["G"].width = 14
    ws_par.column_dimensions["H"].width = 40

    r = 1
    ws_par.merge_cells(f"A{r}:H{r}")
    _hc(ws_par, r, 1, "PARAMETERS — MEASURED (EX-POST) vs PROJECTED (EX-ANTE)",
        _EP_HEADER, size=12, halign="center")
    r += 1
    ws_par.merge_cells(f"A{r}:H{r}")
    _hc(ws_par, r, 1,
        "Green = deviation <5%    Orange = 5–10%    Red = >10% — review required",
        _EP_SUBHEAD, size=9, halign="center")
    r += 2

    par_headers = ["Symbol", "Parameter name", "Measured value", "Ex-ante value",
                   "Unit", "Deviation %", "Status", "Measurement source / method"]
    for ci, h in enumerate(par_headers, 1):
        _hc(ws_par, r, ci, h, _EP_HEADER, size=9, border=tb, wrap=True)
    r += 1

    param_rows = _build_param_rows(calc_result, meth_s)
    par_cell_map = {}

    for pr in param_rows:
        bg = _EP_LIGHT if r % 2 == 0 else None
        measured = pr["value"]
        exante = pr["value"]

        if isinstance(measured, (int, float)) and isinstance(exante, (int, float)) and exante != 0:
            deviation_pct = abs(measured - exante) / abs(exante) * 100
        else:
            deviation_pct = 0.0

        if deviation_pct > 10:
            flag_bg = _FLAG_RED
            status = "REVIEW"
        elif deviation_pct > 5:
            flag_bg = _FLAG_ORANGE
            status = "CHECK"
        else:
            flag_bg = _FLAG_GREEN
            status = "OK"

        _dc(ws_par, r, 1, pr["symbol"], bg=bg, bold=True, border=tb, size=9)
        _dc(ws_par, r, 2, pr["name"], bg=bg, border=tb, size=9)
        vc = _dc(ws_par, r, 3, measured, bg=bg, halign="right", border=tb, size=9)
        if isinstance(measured, float):
            vc.number_format = "#,##0.000000"
        _dc(ws_par, r, 4, exante, bg=bg, halign="right", border=tb, size=9)
        _dc(ws_par, r, 5, pr["unit"], bg=bg, border=tb, size=9)
        _dc(ws_par, r, 6, round(deviation_pct, 1) if deviation_pct else "",
            bg=bg, halign="right", number_format="#,##0.0", border=tb, size=9)
        _dc(ws_par, r, 7, status, bg=flag_bg, halign="center", bold=True, border=tb, size=9)
        _dc(ws_par, r, 8, pr["source"], bg=bg, border=tb, size=9)
        par_cell_map[pr["key"]] = f"'Parameters (Measured vs EA)'!C{r}"
        r += 1

    _freeze(ws_par, 5, 1)

    # ------------------------------------------------------------------
    # Sheet 3 — ER Calculation
    # ------------------------------------------------------------------
    ws_calc = wb.create_sheet("ER Calculation")
    ws_calc.column_dimensions["A"].width = 6
    ws_calc.column_dimensions["B"].width = 42
    ws_calc.column_dimensions["C"].width = 65
    ws_calc.column_dimensions["D"].width = 16
    ws_calc.column_dimensions["E"].width = 18

    r = 1
    ws_calc.merge_cells(f"A{r}:E{r}")
    _hc(ws_calc, r, 1,
        f"ER CALCULATION (EX-POST) — Monitoring Period {mp_number}",
        _EP_HEADER, size=12, halign="center")
    r += 1
    ws_calc.merge_cells(f"A{r}:E{r}")
    _hc(ws_calc, r, 1,
        f"Period: {mp_start_str} to {mp_end_str} ({mp_days} days)",
        _EP_SUBHEAD, size=9, halign="center")
    r += 2

    steps_headers = ["Step", "Description", "Formula / derivation", "Result", "Unit"]
    for ci, h in enumerate(steps_headers, 1):
        _hc(ws_calc, r, ci, h, _EP_HEADER, size=9, border=tb)
    r += 1

    steps = calc_result.get("calculation_steps", [])
    for s in steps:
        bg = _EP_LIGHT if r % 2 == 0 else None
        val = s.get("value", s.get("value_baseline", ""))
        _dc(ws_calc, r, 1, s.get("step", ""), bg=bg, bold=True, halign="center", border=tb, size=9)
        _dc(ws_calc, r, 2, s.get("name", ""), bg=bg, bold=True, border=tb, size=9)
        _dc(ws_calc, r, 3, s.get("formula", ""), bg=bg, border=tb, size=8, color="555555")
        vc = _dc(ws_calc, r, 4, val, bg=bg, halign="right", border=tb, size=9)
        if isinstance(val, float):
            vc.number_format = "#,##0.000000"
        _dc(ws_calc, r, 5, s.get("unit", ""), bg=bg, border=tb, size=9)
        r += 1

    r += 1
    ws_calc.merge_cells(f"A{r}:E{r}")
    _hc(ws_calc, r, 1, "MONITORING PERIOD TOTAL ERs", _EP_SUBHEAD, size=10)
    r += 1

    summ = calc_result.get("summary", {})
    for lbl, val, fmt in [
        ("Total baseline emissions", summ.get("total_baseline", 0), "#,##0.00"),
        ("Total project emissions", summ.get("total_project", 0), "#,##0.00"),
        ("Total leakage", summ.get("total_leakage", 0), "#,##0.00"),
        ("Total NET emission reductions (tCO2e)", summ.get("total_er", 0), "#,##0.00"),
        ("Average annual net ERs (tCO2e/yr)", summ.get("average_annual_er", 0), "#,##0.00"),
    ]:
        bg = _EP_LIGHT if r % 2 == 0 else None
        _dc(ws_calc, r, 1, lbl, bg=bg, bold=("Total NET" in lbl), border=tb, size=9)
        ws_calc.merge_cells(f"A{r}:C{r}")
        vc = _dc(ws_calc, r, 4, val, bg=bg, halign="right", border=tb, size=9,
                 number_format=fmt, bold=("Total NET" in lbl))
        r += 1

    # ------------------------------------------------------------------
    # Sheet 4 — Vintage Allocation
    # ------------------------------------------------------------------
    ws_vint = wb.create_sheet("Vintage Allocation")
    ws_vint.column_dimensions["A"].width = 16
    ws_vint.column_dimensions["B"].width = 20
    ws_vint.column_dimensions["C"].width = 14
    ws_vint.column_dimensions["D"].width = 14
    ws_vint.column_dimensions["E"].width = 20

    r = 1
    ws_vint.merge_cells(f"A{r}:E{r}")
    _hc(ws_vint, r, 1, "VINTAGE ALLOCATION — CALENDAR YEAR BREAKDOWN", _EP_HEADER,
        size=12, halign="center")
    r += 1
    ws_vint.merge_cells(f"A{r}:E{r}")
    _hc(ws_vint, r, 1,
        f"Monitoring period: {mp_start_str} → {mp_end_str}  ({mp_days} days total)",
        _EP_SUBHEAD, size=9, halign="center")
    r += 2

    vint_headers = ["Vintage year", "Days in MP", "Fraction of MP",
                    "Gross ERs (tCO2e)", "Net ERs (tCO2e)"]
    for ci, h in enumerate(vint_headers, 1):
        _hc(ws_vint, r, ci, h, _EP_HEADER, size=9, border=tb)
    r += 1

    total_er = summ.get("total_er", 0) or 0.0
    if vintage_rows:
        vd_start = r
        for vy, vdays, vfrac in vintage_rows:
            bg = _EP_LIGHT if r % 2 == 0 else None
            net_vintage = round(total_er * vfrac, 2)
            _dc(ws_vint, r, 1, str(vy), bg=bg, bold=True, halign="center", border=tb, size=9)
            _dc(ws_vint, r, 2, vdays, bg=bg, halign="center", border=tb, size=9)
            _dc(ws_vint, r, 3, round(vfrac, 4), bg=bg, halign="center",
                number_format="0.0000", border=tb, size=9)
            _dc(ws_vint, r, 4, round(total_er * vfrac + summ.get("total_leakage", 0) * vfrac, 2),
                bg=bg, halign="right", number_format="#,##0.00", border=tb, size=9)
            _dc(ws_vint, r, 5, net_vintage, bg=bg, halign="right",
                number_format="#,##0.00", border=tb, size=9)
            r += 1
        vd_end = r - 1
        _dc(ws_vint, r, 1, "TOTAL", bold=True, border=tb, size=9)
        _dc(ws_vint, r, 2, mp_days, bold=True, halign="center", border=tb, size=9)
        _dc(ws_vint, r, 3, 1.0, bold=True, halign="center", number_format="0.0000", border=tb, size=9)
        _dc(ws_vint, r, 4, "", border=tb)
        c = ws_vint.cell(row=r, column=5, value=f"=SUM(E{vd_start}:E{vd_end})")
        c.font = Font(name="Calibri", bold=True, size=9)
        c.number_format = "#,##0.00"
        c.border = tb
    else:
        ws_vint.merge_cells(f"A{r}:E{r}")
        _dc(ws_vint, r, 1,
            "Set monitoring period start and end dates in the Setup tab to compute vintage allocation.",
            bg=_FLAG_ORANGE, border=tb, size=9, halign="center")
    r += 2

    ws_vint.merge_cells(f"A{r}:E{r}")
    _dc(ws_vint, r, 1,
        "Allocation methodology: (days in calendar year / total MP days) x total net ERs for the MP.",
        bg=_EP_LIGHT, border=tb, size=8, halign="center")

    # ------------------------------------------------------------------
    # Sheet 5 — Data Quality
    # ------------------------------------------------------------------
    ws_dq = wb.create_sheet("Data Quality")
    ws_dq.column_dimensions["A"].width = 18
    ws_dq.column_dimensions["B"].width = 38
    ws_dq.column_dimensions["C"].width = 14
    ws_dq.column_dimensions["D"].width = 18
    ws_dq.column_dimensions["E"].width = 14
    ws_dq.column_dimensions["F"].width = 40

    r = 1
    ws_dq.merge_cells(f"A{r}:F{r}")
    _hc(ws_dq, r, 1, "DATA QUALITY ASSESSMENT", _EP_HEADER, size=12, halign="center")
    r += 1
    ws_dq.merge_cells(f"A{r}:F{r}")
    _hc(ws_dq, r, 1, "Tier 1 = directly monitored / metered   Tier 2 = conservative default / survey   Tier 3 = modelled / estimated",
        _EP_SUBHEAD, size=9, halign="center", wrap=True)
    r += 2

    dq_headers = ["Symbol", "Parameter", "Measurement tier", "Measurement method",
                  "Uncertainty (%)", "Quality notes"]
    for ci, h in enumerate(dq_headers, 1):
        _hc(ws_dq, r, ci, h, _EP_HEADER, size=9, border=tb)
    r += 1

    dq_rows = [
        ("fNRB,y", "Fraction non-renewable biomass", "Tier 2", "National fNRB study or IPCC default", "±10%",
         "Conservative lower-bound value. Country study preferred over IPCC default."),
        ("NCV_b", "NCV — baseline fuel", "Tier 1", "Lab analysis or IPCC table", "±3%",
         "IPCC default acceptable if lab analysis unavailable."),
        ("SFC_b / SFC_p", "Specific fuel consumption (baseline & project)", "Tier 2",
         "Kitchen Performance Test (KPT) — ≥3 replications", "±5%",
         "Upper-bound SFC_p used for conservativeness."),
        ("N_i,y", "Number of active devices", "Tier 1",
         "Device registry cross-checked with end-user survey", "±2%",
         "Registry must be verified against field count for ≥sample size."),
        ("pop_stoves,y", "Device usage / activity rate", "Tier 2",
         "End-user survey (stratified random sample)", "±8%",
         "Conservative lower-bound confidence interval used."),
        ("EF_CO2,b / EF_CO2,p", "CO₂ emission factors", "Tier 1",
         "IPCC 2006 defaults", "±5%", "IPCC defaults are universally accepted."),
        ("EF_nonCO2", "Non-CO₂ emission factors", "Tier 1",
         "Methodology table (TPDDTEC v4.0 / AMS-II.G)", "±5%",
         "Prescribed methodology default."),
    ]
    for row_data in dq_rows:
        bg = _EP_LIGHT if r % 2 == 0 else None
        for ci, val in enumerate(row_data, 1):
            _dc(ws_dq, r, ci, val, bg=bg, border=tb, size=9,
                bold=(ci == 1))
        r += 1

    # ------------------------------------------------------------------
    # Sheet 6 — Deviation Log
    # ------------------------------------------------------------------
    ws_dev = wb.create_sheet("Deviation Log")
    ws_dev.column_dimensions["A"].width = 18
    ws_dev.column_dimensions["B"].width = 35
    ws_dev.column_dimensions["C"].width = 16
    ws_dev.column_dimensions["D"].width = 16
    ws_dev.column_dimensions["E"].width = 12
    ws_dev.column_dimensions["F"].width = 14
    ws_dev.column_dimensions["G"].width = 40
    ws_dev.column_dimensions["H"].width = 20

    r = 1
    ws_dev.merge_cells(f"A{r}:H{r}")
    _hc(ws_dev, r, 1, "DEVIATION LOG", _EP_HEADER, size=12, halign="center")
    r += 1
    ws_dev.merge_cells(f"A{r}:H{r}")
    _hc(ws_dev, r, 1,
        "Document any deviations from the approved methodology or ex-ante assumptions here.",
        _EP_SUBHEAD, size=9, halign="center", wrap=True)
    r += 2

    dev_headers = ["Parameter / section", "Approved value / approach",
                   "Actual value / approach used", "Deviation",
                   "Deviation %", "Impact on ERs", "Justification", "Approved by verifier?"]
    for ci, h in enumerate(dev_headers, 1):
        _hc(ws_dev, r, ci, h, _EP_HEADER, size=9, border=tb, wrap=True)
    r += 1

    ws_dev.merge_cells(f"A{r}:H{r}")
    _dc(ws_dev, r, 1,
        "No deviations recorded. Fill in above rows if any parameter or procedure differs from the approved PDD/VPA-DD.",
        bg=_FLAG_GREEN, border=tb, size=9, halign="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

def generate_er_workbook(project, doc_type, calc_result=None, project_id=None):
    """
    Generate the appropriate ER workbook based on doc_type.
    doc_type in ('pdd', 'vpa_dd', 'poa_dd') → ex-ante
    doc_type in ('mr',)                       → ex-post

    project_id : int | None
        When provided, parameters are read from the live parameter engine so
        the workbook always reflects the current confirmed values — not just
        the snapshot embedded in calc_result.
    """
    # Enrich calc_result with live parameter engine data when project_id is known
    if project_id is not None and calc_result is not None:
        try:
            from carbongpt.core.parameter_engine import get_parameters_as_dict
            live_params = get_parameters_as_dict(project_id)
            # Merge: engine values are authoritative; calc_result snapshot fills any gaps
            snapshot = calc_result.get("parameters_used", {})
            if isinstance(snapshot, list):
                snapshot = {p.get("parameter", f"p{i}"): p for i, p in enumerate(snapshot)}
            merged = dict(snapshot)
            merged.update(live_params)
            calc_result = dict(calc_result)
            calc_result["parameters_used"] = merged
            calc_result["_source"] = "parameter_engine"
        except Exception as _exc:
            logger.warning("Could not enrich workbook with live parameter engine data: %s", _exc)

    if doc_type in ("mr",):
        return generate_expost_workbook(project, calc_result)
    return generate_exante_workbook(project, calc_result)

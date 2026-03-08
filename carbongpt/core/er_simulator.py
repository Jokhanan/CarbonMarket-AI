import logging
import json
import math
from carbongpt.repository.db import get_cursor
from carbongpt.core.parameter_engine import get_parameters_as_dict

logger = logging.getLogger(__name__)


def calculate_cookstove_er(params, crediting_years=7, start_year=2025, methodology="VM0050"):
    fNRB = _pval(params, "fNRB", 0.30)
    NCV_b = _pval(params, "NCV_baseline", 15.6)
    EF_CO2_b = _pval(params, "EF_CO2_baseline", 112.0)
    EF_nonCO2_b = _pval(params, "EF_nonCO2_baseline", 9.46)
    NCV_p = _pval(params, "NCV_project", NCV_b)
    EF_CO2_p = _pval(params, "EF_CO2_project", EF_CO2_b)
    EF_nonCO2_p = _pval(params, "EF_nonCO2_project", EF_nonCO2_b)
    CF = _pval(params, "CF", 1.0)
    leakage_pct = 1.0 - _pval(params, "leakage_discount", 0.95)
    usage_rate_base = _pval(params, "usage_rate", 0.90)
    num_hh = _pval(params, "num_households", 1000)
    hh_size = _pval(params, "household_size", 5.0)

    baseline_fuel = _ptext(params, "baseline_fuel", "wood")
    is_charcoal_baseline = baseline_fuel.lower() in ("charcoal", "charbon")

    if methodology in ("VM0050",):
        bl_consumption = _pval(params, "baseline_fuel_consumption", hh_size * 0.4)
        pj_consumption = _pval(params, "project_fuel_consumption", bl_consumption * 0.5)
        consumption_label = "baseline_fuel_consumption"
        consumption_pj_label = "project_fuel_consumption"
    else:
        sfc_b = _pval(params, "SFC_baseline", 400.0)
        sfc_p = _pval(params, "SFC_project", 200.0)
        bl_consumption = sfc_b * hh_size / 1000.0
        pj_consumption = sfc_p * hh_size / 1000.0
        consumption_label = f"SFC_baseline * household_size / 1000 = {sfc_b} * {hh_size} / 1000"
        consumption_pj_label = f"SFC_project * household_size / 1000 = {sfc_p} * {hh_size} / 1000"

    if is_charcoal_baseline and CF > 1.0:
        bl_consumption_wood_equiv = bl_consumption * CF
        pj_consumption_wood_equiv = pj_consumption * CF
        cf_note = f"Charcoal baseline: consumption * CF = {bl_consumption:.4f} * {CF} = {bl_consumption_wood_equiv:.4f} t wood-equiv/hh/yr"
    else:
        bl_consumption_wood_equiv = bl_consumption
        pj_consumption_wood_equiv = pj_consumption
        cf_note = f"Wood baseline: CF not applied (CF={CF})"

    ec_b = bl_consumption_wood_equiv * NCV_b / 1000.0
    be_per_hh = ec_b * (EF_CO2_b + EF_nonCO2_b) * fNRB
    ec_p = pj_consumption_wood_equiv * NCV_p / 1000.0
    pe_per_hh = ec_p * (EF_CO2_p + EF_nonCO2_p) * fNRB

    calculation_steps = [
        {
            "step": 1,
            "name": "Baseline fuel consumption per household",
            "formula": consumption_label,
            "value": round(bl_consumption, 6),
            "unit": "t/hh/yr",
        },
        {
            "step": 2,
            "name": "Project fuel consumption per household",
            "formula": consumption_pj_label,
            "value": round(pj_consumption, 6),
            "unit": "t/hh/yr",
        },
        {
            "step": 3,
            "name": "CF adjustment (charcoal-to-wood equivalent)",
            "formula": cf_note,
            "value_baseline": round(bl_consumption_wood_equiv, 6),
            "value_project": round(pj_consumption_wood_equiv, 6),
            "unit": "t wood-equiv/hh/yr",
        },
        {
            "step": 4,
            "name": "Baseline energy content per household",
            "formula": f"B_cons_wood_equiv * NCV_b / 1000 = {bl_consumption_wood_equiv:.6f} * {NCV_b} / 1000",
            "value": round(ec_b, 6),
            "unit": "TJ/hh/yr",
        },
        {
            "step": 5,
            "name": "Baseline emissions per household",
            "formula": f"EC_b * (EF_CO2_b + EF_nonCO2_b) * fNRB = {ec_b:.6f} * ({EF_CO2_b} + {EF_nonCO2_b}) * {fNRB}",
            "value": round(be_per_hh, 6),
            "unit": "tCO2e/hh/yr",
        },
        {
            "step": 6,
            "name": "Project energy content per household",
            "formula": f"P_cons_wood_equiv * NCV_p / 1000 = {pj_consumption_wood_equiv:.6f} * {NCV_p} / 1000",
            "value": round(ec_p, 6),
            "unit": "TJ/hh/yr",
        },
        {
            "step": 7,
            "name": "Project emissions per household",
            "formula": f"EC_p * (EF_CO2_p + EF_nonCO2_p) * fNRB = {ec_p:.6f} * ({EF_CO2_p} + {EF_nonCO2_p}) * {fNRB}",
            "value": round(pe_per_hh, 6),
            "unit": "tCO2e/hh/yr",
        },
        {
            "step": 8,
            "name": "ER per household per year (before leakage)",
            "formula": f"BE_per_hh - PE_per_hh = {be_per_hh:.6f} - {pe_per_hh:.6f}",
            "value": round(be_per_hh - pe_per_hh, 6),
            "unit": "tCO2e/hh/yr",
        },
    ]

    years = []
    total_er = 0.0
    total_be = 0.0
    total_pe = 0.0
    total_le = 0.0

    for y in range(crediting_years):
        year_num = y + 1
        cal_year = start_year + y

        usage_rate = max(usage_rate_base - (y * 0.02), 0.50)
        active_hh = num_hh * usage_rate
        be_y = be_per_hh * active_hh
        pe_y = pe_per_hh * active_hh
        gross_er = be_y - pe_y
        le_y = gross_er * leakage_pct if leakage_pct > 0 else 0.0
        er_y = gross_er - le_y

        total_er += er_y
        total_be += be_y
        total_pe += pe_y
        total_le += le_y

        years.append({
            "year_number": year_num,
            "calendar_year": cal_year,
            "usage_rate": round(usage_rate, 4),
            "active_households": round(active_hh, 2),
            "active_hh_formula": f"max({usage_rate_base} - ({y} * 0.02), 0.50) * {num_hh}",
            "be_per_hh": round(be_per_hh, 6),
            "pe_per_hh": round(pe_per_hh, 6),
            "baseline_emissions": round(be_y, 2),
            "baseline_formula": f"{be_per_hh:.4f} * {active_hh:.0f}",
            "project_emissions": round(pe_y, 2),
            "project_formula": f"{pe_per_hh:.4f} * {active_hh:.0f}",
            "gross_er": round(gross_er, 2),
            "leakage": round(le_y, 2),
            "leakage_formula": f"{gross_er:.2f} * {leakage_pct:.4f}" if leakage_pct > 0 else "0",
            "net_er": round(er_y, 2),
            "net_er_formula": f"{gross_er:.2f} - {le_y:.2f}",
        })

    return {
        "calculation_steps": calculation_steps,
        "years": years,
        "summary": {
            "total_er": round(total_er, 2),
            "total_baseline": round(total_be, 2),
            "total_project": round(total_pe, 2),
            "total_leakage": round(total_le, 2),
            "average_annual_er": round(total_er / crediting_years, 2),
            "crediting_years": crediting_years,
            "methodology": methodology,
        },
        "parameters_used": {
            "fNRB": {"value": fNRB, "unit": "fraction", "description": "Fraction of non-renewable biomass"},
            "NCV_baseline": {"value": NCV_b, "unit": "TJ/Gg", "description": "Net calorific value (baseline fuel)"},
            "NCV_project": {"value": NCV_p, "unit": "TJ/Gg", "description": "Net calorific value (project fuel)"},
            "EF_CO2_baseline": {"value": EF_CO2_b, "unit": "tCO2/TJ", "description": "CO2 emission factor (baseline)"},
            "EF_nonCO2_baseline": {"value": EF_nonCO2_b, "unit": "tCO2e/TJ", "description": "Non-CO2 emission factor (baseline)"},
            "EF_CO2_project": {"value": EF_CO2_p, "unit": "tCO2/TJ", "description": "CO2 emission factor (project)"},
            "EF_nonCO2_project": {"value": EF_nonCO2_p, "unit": "tCO2e/TJ", "description": "Non-CO2 emission factor (project)"},
            "CF": {"value": CF, "unit": "kg wood/kg charcoal", "description": "Charcoal-to-wood conversion factor"},
            "baseline_fuel": {"value": baseline_fuel, "unit": "", "description": "Baseline fuel type"},
            "is_charcoal_baseline": {"value": is_charcoal_baseline, "unit": "", "description": "Charcoal baseline applied"},
            "bl_consumption": {"value": round(bl_consumption, 4), "unit": "t/hh/yr", "description": "Baseline consumption (raw)"},
            "bl_consumption_wood_equiv": {"value": round(bl_consumption_wood_equiv, 4), "unit": "t/hh/yr", "description": "Baseline consumption (wood-equiv)"},
            "pj_consumption": {"value": round(pj_consumption, 4), "unit": "t/hh/yr", "description": "Project consumption (raw)"},
            "pj_consumption_wood_equiv": {"value": round(pj_consumption_wood_equiv, 4), "unit": "t/hh/yr", "description": "Project consumption (wood-equiv)"},
            "leakage_pct": {"value": leakage_pct, "unit": "fraction", "description": "Leakage deduction percentage"},
            "usage_rate": {"value": usage_rate_base, "unit": "fraction", "description": "Initial usage rate (decays 2%/yr)"},
            "num_households": {"value": num_hh, "unit": "count", "description": "Number of households"},
            "household_size": {"value": hh_size, "unit": "persons", "description": "Average household size"},
        },
    }


def calculate_cookstove_er_cohort(params, crediting_years=7, start_year=2025, methodology="VM0050",
                                   deployment_config=None):
    if deployment_config is None:
        deployment_config = {}

    fNRB = _pval(params, "fNRB", 0.30)
    NCV_b = _pval(params, "NCV_baseline", 15.6)
    EF_CO2_b = _pval(params, "EF_CO2_baseline", 112.0)
    EF_nonCO2_b = _pval(params, "EF_nonCO2_baseline", 9.46)
    NCV_p = _pval(params, "NCV_project", NCV_b)
    EF_CO2_p = _pval(params, "EF_CO2_project", EF_CO2_b)
    EF_nonCO2_p = _pval(params, "EF_nonCO2_project", EF_nonCO2_b)
    CF = _pval(params, "CF", 1.0)
    leakage_pct = 1.0 - _pval(params, "leakage_discount", 0.95)
    num_hh = int(_pval(params, "num_households", 1000))
    hh_size = _pval(params, "household_size", 5.0)

    baseline_fuel = _ptext(params, "baseline_fuel", "wood")
    is_charcoal_baseline = baseline_fuel.lower() in ("charcoal", "charbon")

    if methodology in ("VM0050",):
        bl_consumption = _pval(params, "baseline_fuel_consumption", hh_size * 0.4)
        pj_consumption = _pval(params, "project_fuel_consumption", bl_consumption * 0.5)
    else:
        sfc_b = _pval(params, "SFC_baseline", 400.0)
        sfc_p = _pval(params, "SFC_project", 200.0)
        bl_consumption = sfc_b * hh_size / 1000.0
        pj_consumption = sfc_p * hh_size / 1000.0

    if is_charcoal_baseline and CF > 1.0:
        bl_consumption_wood_equiv = bl_consumption * CF
        pj_consumption_wood_equiv = pj_consumption * CF
    else:
        bl_consumption_wood_equiv = bl_consumption
        pj_consumption_wood_equiv = pj_consumption

    ec_b = bl_consumption_wood_equiv * NCV_b / 1000.0
    be_per_unit = ec_b * (EF_CO2_b + EF_nonCO2_b) * fNRB
    ec_p = pj_consumption_wood_equiv * NCV_p / 1000.0
    pe_per_unit = ec_p * (EF_CO2_p + EF_nonCO2_p) * fNRB
    er_per_unit = be_per_unit - pe_per_unit

    deployment_mode = deployment_config.get("deployment_mode", "instant")
    monthly_deployment = deployment_config.get("monthly_deployment", 500)
    custom_schedule = deployment_config.get("custom_schedule", [])
    deployment_timing = deployment_config.get("deployment_timing", "mid")
    tech_lifetime_years = deployment_config.get("tech_lifetime_years", 5.0)
    dropoff_mode = deployment_config.get("dropoff_mode", "annual_rate")
    annual_dropoff_rate = deployment_config.get("annual_dropoff_rate", 0.10)
    custom_dropoff_curve = deployment_config.get("custom_dropoff_curve", [])
    usage_rate_mode = deployment_config.get("usage_rate_mode", "fixed")
    usage_rate_fixed = deployment_config.get("usage_rate", 0.90)
    usage_curve = deployment_config.get("usage_curve", [])

    timing_factor_map = {"start": 1.0, "mid": 0.5, "end": 0.0}
    timing_factor = timing_factor_map.get(deployment_timing, 0.5)

    total_months = crediting_years * 12
    cohorts = []

    if deployment_mode == "instant":
        cohorts.append({"month": 0, "count": num_hh})
    elif deployment_mode == "fixed_monthly":
        deployed_so_far = 0
        month = 0
        while deployed_so_far < num_hh and month < total_months:
            batch = min(monthly_deployment, num_hh - deployed_so_far)
            if batch > 0:
                cohorts.append({"month": month, "count": batch})
                deployed_so_far += batch
            month += 1
    elif deployment_mode == "custom":
        if custom_schedule:
            for entry in custom_schedule:
                m = int(entry.get("month", 0))
                c = int(entry.get("count", 0))
                if c > 0 and m < total_months:
                    cohorts.append({"month": m, "count": c})
        if not cohorts:
            cohorts.append({"month": 0, "count": num_hh})

    def _get_survival(age_years):
        if age_years <= 0:
            return 1.0
        if age_years > tech_lifetime_years:
            return 0.0
        if dropoff_mode == "custom_curve" and custom_dropoff_curve:
            sorted_curve = sorted(custom_dropoff_curve, key=lambda x: x.get("year", 0))
            for i, point in enumerate(sorted_curve):
                py = point.get("year", 0)
                pv = point.get("survival_fraction", 1.0)
                if age_years <= py:
                    if i == 0:
                        prev_y, prev_v = 0, 1.0
                    else:
                        prev_y = sorted_curve[i - 1].get("year", 0)
                        prev_v = sorted_curve[i - 1].get("survival_fraction", 1.0)
                    if py == prev_y:
                        return pv
                    frac = (age_years - prev_y) / (py - prev_y)
                    return prev_v + frac * (pv - prev_v)
            return sorted_curve[-1].get("survival_fraction", 0.0)
        return (1.0 - annual_dropoff_rate) ** age_years

    def _get_usage(age_years):
        if usage_rate_mode == "curve" and usage_curve:
            sorted_uc = sorted(usage_curve, key=lambda x: x.get("year", 0))
            for i, point in enumerate(sorted_uc):
                py = point.get("year", 0)
                pv = point.get("rate", 0.9)
                if age_years <= py:
                    if i == 0:
                        return pv
                    prev_y = sorted_uc[i - 1].get("year", 0)
                    prev_v = sorted_uc[i - 1].get("rate", 0.9)
                    if py == prev_y:
                        return pv
                    frac = (age_years - prev_y) / (py - prev_y)
                    return prev_v + frac * (pv - prev_v)
            return sorted_uc[-1].get("rate", 0.5)
        return usage_rate_fixed

    years = []
    total_er = 0.0
    total_be = 0.0
    total_pe = 0.0
    total_le = 0.0
    cumulative_er = 0.0
    deployment_timeline = []

    for y in range(crediting_years):
        year_num = y + 1
        cal_year = start_year + y
        year_start_month = y * 12
        year_end_month = (y + 1) * 12

        deployed_this_year = 0
        active_units = 0.0
        surviving_units = 0.0
        effectively_used = 0.0
        be_y = 0.0
        pe_y = 0.0

        for cohort in cohorts:
            cm = cohort["month"]
            cc = cohort["count"]

            if cm >= year_end_month:
                continue

            if cm >= year_start_month:
                deployed_this_year += cc
                months_active_in_year = year_end_month - cm
                if deployment_mode == "instant" and cm == 0:
                    fraction_of_year = 1.0
                else:
                    fraction_of_year = (months_active_in_year / 12.0) * timing_factor
                    fraction_of_year = max(fraction_of_year, 0.0)
            else:
                fraction_of_year = 1.0

            age_mid = (year_start_month + 6 - cm) / 12.0
            if cm >= year_start_month:
                age_mid = (cm + (year_end_month - cm) / 2.0 - cm) / 12.0
                age_mid = max((year_end_month - cm) / 2.0 / 12.0, 0.01)

            age_end = (year_end_month - cm) / 12.0

            if age_end > tech_lifetime_years:
                if age_end - 1.0 >= tech_lifetime_years:
                    continue
                fraction_alive = max(tech_lifetime_years - (age_end - 1.0), 0) / 1.0
                fraction_of_year *= fraction_alive

            survival = _get_survival(age_mid)
            surviving_count = cc * survival
            usage = _get_usage(age_mid)
            effective = surviving_count * usage

            active_units += surviving_count
            surviving_units += surviving_count
            effectively_used += effective * fraction_of_year

            cohort_be = be_per_unit * effective * fraction_of_year
            cohort_pe = pe_per_unit * effective * fraction_of_year
            be_y += cohort_be
            pe_y += cohort_pe

        gross_er = be_y - pe_y
        le_y = gross_er * leakage_pct if leakage_pct > 0 else 0.0
        er_y = gross_er - le_y

        total_er += er_y
        total_be += be_y
        total_pe += pe_y
        total_le += le_y
        cumulative_er += er_y

        cumulative_deployed = sum(c["count"] for c in cohorts if c["month"] < year_end_month)

        years.append({
            "year_number": year_num,
            "calendar_year": cal_year,
            "deployed_this_year": deployed_this_year,
            "cumulative_deployed": cumulative_deployed,
            "active_units": round(active_units, 0),
            "surviving_units": round(surviving_units, 0),
            "effectively_used": round(effectively_used, 0),
            "baseline_emissions": round(be_y, 2),
            "project_emissions": round(pe_y, 2),
            "gross_er": round(gross_er, 2),
            "leakage": round(le_y, 2),
            "net_er": round(er_y, 2),
            "cumulative_er": round(cumulative_er, 2),
        })

        deployment_timeline.append({
            "year": cal_year,
            "year_number": year_num,
            "deployed": deployed_this_year,
            "cumulative_deployed": cumulative_deployed,
            "active": round(active_units, 0),
            "surviving": round(surviving_units, 0),
            "effectively_used": round(effectively_used, 0),
            "net_er": round(er_y, 2),
            "cumulative_er": round(cumulative_er, 2),
        })

    calculation_steps = [
        {"step": 1, "name": "ER per unit per year (before leakage)", "formula": f"BE_unit - PE_unit = {be_per_unit:.6f} - {pe_per_unit:.6f}", "value": round(er_per_unit, 6), "unit": "tCO2e/unit/yr"},
        {"step": 2, "name": "Deployment mode", "formula": deployment_mode, "value": deployment_mode, "unit": ""},
        {"step": 3, "name": "Total units to deploy", "formula": f"num_households = {num_hh}", "value": num_hh, "unit": "units"},
        {"step": 4, "name": "Technology lifetime", "formula": f"{tech_lifetime_years} years", "value": tech_lifetime_years, "unit": "years"},
        {"step": 5, "name": "Drop-off model", "formula": f"{dropoff_mode}: {annual_dropoff_rate*100:.1f}%/yr" if dropoff_mode == "annual_rate" else f"{dropoff_mode}: custom curve", "value": annual_dropoff_rate if dropoff_mode == "annual_rate" else "custom", "unit": ""},
        {"step": 6, "name": "Usage rate", "formula": f"{usage_rate_mode}: {usage_rate_fixed:.0%}" if usage_rate_mode == "fixed" else f"{usage_rate_mode}: custom curve", "value": usage_rate_fixed if usage_rate_mode == "fixed" else "custom", "unit": ""},
        {"step": 7, "name": "Deployment timing", "formula": f"{deployment_timing} of period (factor: {timing_factor})", "value": timing_factor, "unit": ""},
    ]

    return {
        "calculation_steps": calculation_steps,
        "years": years,
        "deployment_timeline": deployment_timeline,
        "cohort_count": len(cohorts),
        "cohorts_summary": [{"month": c["month"], "count": c["count"]} for c in cohorts[:50]],
        "summary": {
            "total_er": round(total_er, 2),
            "total_baseline": round(total_be, 2),
            "total_project": round(total_pe, 2),
            "total_leakage": round(total_le, 2),
            "average_annual_er": round(total_er / crediting_years, 2) if crediting_years > 0 else 0,
            "crediting_years": crediting_years,
            "methodology": methodology,
            "deployment_mode": deployment_mode,
            "tech_lifetime_years": tech_lifetime_years,
            "peak_active_units": max((y["active_units"] for y in years), default=0),
            "peak_surviving_units": max((y["surviving_units"] for y in years), default=0),
        },
        "parameters_used": {
            "fNRB": {"value": fNRB, "unit": "fraction", "description": "Fraction of non-renewable biomass"},
            "NCV_baseline": {"value": NCV_b, "unit": "TJ/Gg", "description": "Net calorific value (baseline fuel)"},
            "EF_CO2_baseline": {"value": EF_CO2_b, "unit": "tCO2/TJ", "description": "CO2 emission factor (baseline)"},
            "EF_nonCO2_baseline": {"value": EF_nonCO2_b, "unit": "tCO2e/TJ", "description": "Non-CO2 emission factor (baseline)"},
            "NCV_project": {"value": NCV_p, "unit": "TJ/Gg", "description": "Net calorific value (project fuel)"},
            "EF_CO2_project": {"value": EF_CO2_p, "unit": "tCO2/TJ", "description": "CO2 emission factor (project)"},
            "EF_nonCO2_project": {"value": EF_nonCO2_p, "unit": "tCO2e/TJ", "description": "Non-CO2 emission factor (project)"},
            "CF": {"value": CF, "unit": "kg wood/kg charcoal", "description": "Charcoal-to-wood conversion factor"},
            "baseline_fuel": {"value": baseline_fuel, "unit": "", "description": "Baseline fuel type"},
            "num_households": {"value": num_hh, "unit": "count", "description": "Total units to deploy"},
            "leakage_pct": {"value": leakage_pct, "unit": "fraction", "description": "Leakage deduction percentage"},
            "deployment_mode": {"value": deployment_mode, "unit": "", "description": "Deployment ramp-up mode"},
            "tech_lifetime_years": {"value": tech_lifetime_years, "unit": "years", "description": "Technology operational lifetime"},
            "annual_dropoff_rate": {"value": annual_dropoff_rate, "unit": "fraction", "description": "Annual technology drop-off rate"},
            "usage_rate": {"value": usage_rate_fixed, "unit": "fraction", "description": "Technology usage rate"},
            "deployment_timing": {"value": deployment_timing, "unit": "", "description": "Timing within deployment period"},
        },
    }


def calculate_grid_er(params, crediting_years=7, start_year=2025, methodology="ACM0002"):
    eg_pj = _pval(params, "EG_PJ_y", 50000)
    ef_grid = _pval(params, "EF_grid", 0.8)
    eg_hist = _pval(params, "EG_historical", 0)
    subtype = _ptext(params, "project_subtype", "greenfield")

    if subtype == "greenfield":
        be_formula_desc = f"EG_PJ * EF_grid = {eg_pj:,.0f} * {ef_grid}"
        be_annual = eg_pj * ef_grid
    elif subtype == "capacity_addition":
        net_gen = max(eg_pj - eg_hist, 0)
        be_formula_desc = f"max(EG_PJ - EG_historical, 0) * EF_grid = max({eg_pj:,.0f} - {eg_hist:,.0f}, 0) * {ef_grid} = {net_gen:,.0f} * {ef_grid}"
        be_annual = net_gen * ef_grid
    else:
        be_formula_desc = f"EG_PJ * EF_grid = {eg_pj:,.0f} * {ef_grid}"
        be_annual = eg_pj * ef_grid

    calculation_steps = [
        {
            "step": 1,
            "name": "Net electricity generation (project activity)",
            "formula": f"EG_PJ_y = {eg_pj:,.0f} MWh/yr",
            "value": eg_pj,
            "unit": "MWh/yr",
        },
        {
            "step": 2,
            "name": "Grid emission factor",
            "formula": f"EF_grid = {ef_grid} tCO2/MWh",
            "value": ef_grid,
            "unit": "tCO2/MWh",
        },
    ]

    if subtype == "capacity_addition":
        calculation_steps.append({
            "step": 3,
            "name": "Historical electricity generation (baseline)",
            "formula": f"EG_historical = {eg_hist:,.0f} MWh/yr",
            "value": eg_hist,
            "unit": "MWh/yr",
        })
        calculation_steps.append({
            "step": 4,
            "name": "Baseline emissions (annual)",
            "formula": be_formula_desc,
            "value": round(be_annual, 2),
            "unit": "tCO2e/yr",
        })
    else:
        calculation_steps.append({
            "step": 3,
            "name": "Baseline emissions (annual)",
            "formula": be_formula_desc,
            "value": round(be_annual, 2),
            "unit": "tCO2e/yr",
        })

    calculation_steps.append({
        "step": len(calculation_steps) + 1,
        "name": "Project emissions",
        "formula": "PE_y = 0 (renewable energy, zero direct emissions)",
        "value": 0.0,
        "unit": "tCO2e/yr",
    })
    calculation_steps.append({
        "step": len(calculation_steps) + 1,
        "name": "Leakage",
        "formula": "LE_y = 0 (no leakage for grid-connected renewable energy)",
        "value": 0.0,
        "unit": "tCO2e/yr",
    })
    calculation_steps.append({
        "step": len(calculation_steps) + 1,
        "name": "Net emission reductions (annual)",
        "formula": f"ER_y = BE_y - PE_y - LE_y = {be_annual:,.2f} - 0 - 0",
        "value": round(be_annual, 2),
        "unit": "tCO2e/yr",
    })

    years = []
    total_er = 0.0
    total_be = 0.0

    for y in range(crediting_years):
        year_num = y + 1
        cal_year = start_year + y
        be_y = be_annual
        pe_y = 0.0
        le_y = 0.0
        er_y = be_y - pe_y - le_y
        total_er += er_y
        total_be += be_y

        years.append({
            "year_number": year_num,
            "calendar_year": cal_year,
            "baseline_emissions": round(be_y, 2),
            "baseline_formula": f"{eg_pj:,.0f} * {ef_grid}" if subtype == "greenfield" else f"max({eg_pj:,.0f} - {eg_hist:,.0f}, 0) * {ef_grid}",
            "project_emissions": round(pe_y, 2),
            "project_formula": "0 (renewable)",
            "gross_er": round(er_y, 2),
            "leakage": round(le_y, 2),
            "leakage_formula": "0",
            "net_er": round(er_y, 2),
            "net_er_formula": f"{be_y:,.2f} - 0 - 0",
        })

    return {
        "calculation_steps": calculation_steps,
        "years": years,
        "summary": {
            "total_er": round(total_er, 2),
            "total_baseline": round(total_be, 2),
            "total_project": 0.0,
            "total_leakage": 0.0,
            "average_annual_er": round(total_er / crediting_years, 2),
            "crediting_years": crediting_years,
            "methodology": methodology,
        },
        "parameters_used": {
            "EG_PJ_y": {"value": eg_pj, "unit": "MWh/yr", "description": "Net electricity generation (project)"},
            "EF_grid": {"value": ef_grid, "unit": "tCO2/MWh", "description": "Grid emission factor"},
            "EG_historical": {"value": eg_hist, "unit": "MWh/yr", "description": "Historical electricity generation"},
            "project_subtype": {"value": subtype, "unit": "", "description": "Project subtype (greenfield/capacity_addition)"},
        },
    }


def run_scenario(project_id, scenario_id=None, parameter_overrides=None, deployment_config=None):
    params = get_parameters_as_dict(project_id)

    with get_cursor() as cur:
        cur.execute("SELECT methodology, crediting_period_years, crediting_period_start FROM user_projects WHERE id = %s", (project_id,))
        project = cur.fetchone()
        if not project:
            return {"error": "Project not found"}

    methodology = (project["methodology"] or "").upper().replace("GS-", "")
    crediting_years = project.get("crediting_period_years") or 7
    start_year = 2025
    if project.get("crediting_period_start"):
        start_year = project["crediting_period_start"].year

    if parameter_overrides:
        for key, val in parameter_overrides.items():
            if key in params:
                params[key]["value"] = val
            else:
                params[key] = {"value": val}

    if methodology in ("VM0050", "TPDDTEC"):
        if deployment_config and deployment_config.get("deployment_mode") != "instant_legacy":
            result = calculate_cookstove_er_cohort(params, crediting_years, start_year, methodology, deployment_config)
        else:
            result = calculate_cookstove_er(params, crediting_years, start_year, methodology)
    elif methodology in ("ACM0002", "AMS-I.D.", "AMSID"):
        result = calculate_grid_er(params, crediting_years, start_year, methodology)
    else:
        return {"error": f"ER calculation not implemented for methodology: {methodology}"}

    return result


VALID_SCENARIO_PURPOSES = ("exploratory", "comparison", "shortlisted", "selected_for_drafting", "archived")


def save_scenario(project_id, name, description="", parameter_overrides=None,
                  carbon_price=None, price_escalation=0, developer_share=100,
                  buffer_pool=0, admin_fee=0, is_baseline=False, deployment_config=None,
                  scenario_purpose="exploratory"):
    if scenario_purpose not in VALID_SCENARIO_PURPOSES:
        scenario_purpose = "exploratory"

    result = run_scenario(project_id, parameter_overrides=parameter_overrides, deployment_config=deployment_config)
    if "error" in result:
        return result

    if carbon_price:
        _add_finance(result, carbon_price, price_escalation, developer_share, buffer_pool, admin_fee)

    if is_baseline and scenario_purpose == "exploratory":
        scenario_purpose = "shortlisted"

    with get_cursor() as cur:
        if scenario_purpose == "selected_for_drafting":
            cur.execute("""
                UPDATE er_scenarios SET scenario_purpose = 'shortlisted'
                WHERE project_id = %s AND scenario_purpose = 'selected_for_drafting'
            """, (project_id,))

        cur.execute("SELECT methodology, crediting_period_years FROM user_projects WHERE id = %s", (project_id,))
        project = cur.fetchone()

        cur.execute("""
            INSERT INTO er_scenarios
            (project_id, name, description, is_baseline, parameter_overrides,
             methodology_code, crediting_years, carbon_price_usd,
             price_escalation_pct, developer_share_pct, buffer_pool_pct,
             admin_fee_pct, results_summary, calculated_at, scenario_purpose)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), %s)
            RETURNING id
        """, (
            project_id, name, description, is_baseline,
            json.dumps({**(parameter_overrides or {}), "__deployment_config": deployment_config} if deployment_config else (parameter_overrides or {})),
            project["methodology"],
            project.get("crediting_period_years") or 7,
            carbon_price, price_escalation, developer_share,
            buffer_pool, admin_fee,
            json.dumps(result["summary"]),
            scenario_purpose,
        ))
        scenario_id = cur.fetchone()["id"]

        if scenario_purpose == "selected_for_drafting":
            cur.execute("UPDATE user_projects SET selected_scenario_id = %s WHERE id = %s", (scenario_id, project_id))

        for yr in result["years"]:
            cur.execute("""
                INSERT INTO er_scenario_years
                (scenario_id, year_number, calendar_year, baseline_emissions,
                 project_emissions, leakage, net_er, gross_revenue,
                 deductions, net_revenue, details)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                scenario_id, yr["year_number"], yr["calendar_year"],
                yr["baseline_emissions"], yr["project_emissions"],
                yr["leakage"], yr["net_er"],
                yr.get("gross_revenue", 0), yr.get("deductions", 0),
                yr.get("net_revenue", 0), json.dumps(yr),
            ))

    result["scenario_id"] = scenario_id
    result["scenario_purpose"] = scenario_purpose
    return result


def get_scenarios(project_id):
    with get_cursor() as cur:
        cur.execute("""
            SELECT * FROM er_scenarios
            WHERE project_id = %s ORDER BY created_at DESC
        """, (project_id,))
        return cur.fetchall()


def get_scenario_detail(scenario_id):
    with get_cursor() as cur:
        cur.execute("SELECT * FROM er_scenarios WHERE id = %s", (scenario_id,))
        scenario = cur.fetchone()
        if not scenario:
            return None

        cur.execute("""
            SELECT * FROM er_scenario_years
            WHERE scenario_id = %s ORDER BY year_number
        """, (scenario_id,))
        years = cur.fetchall()

        return {"scenario": scenario, "years": years}


def delete_scenario(scenario_id):
    with get_cursor() as cur:
        cur.execute("SELECT project_id, scenario_purpose FROM er_scenarios WHERE id = %s", (scenario_id,))
        row = cur.fetchone()
        if row and row["scenario_purpose"] == "selected_for_drafting":
            cur.execute("UPDATE user_projects SET selected_scenario_id = NULL WHERE id = %s", (row["project_id"],))
        cur.execute("DELETE FROM er_scenarios WHERE id = %s RETURNING id", (scenario_id,))
        return cur.fetchone()


def select_scenario_for_drafting(project_id, scenario_id):
    with get_cursor() as cur:
        cur.execute("SELECT id FROM er_scenarios WHERE id = %s AND project_id = %s", (scenario_id, project_id))
        if not cur.fetchone():
            return {"error": f"Scenario {scenario_id} not found for project {project_id}"}

        cur.execute("""
            UPDATE er_scenarios
            SET scenario_purpose = 'shortlisted'
            WHERE project_id = %s AND scenario_purpose = 'selected_for_drafting'
        """, (project_id,))

        cur.execute("""
            UPDATE er_scenarios
            SET scenario_purpose = 'selected_for_drafting'
            WHERE id = %s AND project_id = %s
            RETURNING id, name
        """, (scenario_id, project_id))
        updated = cur.fetchone()

        cur.execute("""
            UPDATE user_projects SET selected_scenario_id = %s WHERE id = %s
        """, (scenario_id, project_id))

        return {"selected_scenario_id": updated["id"], "selected_scenario_name": updated["name"]}


def deselect_scenario(project_id):
    with get_cursor() as cur:
        cur.execute("""
            UPDATE er_scenarios
            SET scenario_purpose = 'shortlisted'
            WHERE project_id = %s AND scenario_purpose = 'selected_for_drafting'
            RETURNING id, name
        """, (project_id,))
        demoted = cur.fetchone()

        cur.execute("UPDATE user_projects SET selected_scenario_id = NULL WHERE id = %s", (project_id,))

        if demoted:
            return {"demoted_scenario_id": demoted["id"], "demoted_scenario_name": demoted["name"]}
        return {"message": "No scenario was selected"}


def update_scenario_purpose(project_id, scenario_id, purpose):
    if purpose not in VALID_SCENARIO_PURPOSES:
        return {"error": f"Invalid purpose: {purpose}. Must be one of {VALID_SCENARIO_PURPOSES}"}

    if purpose == "selected_for_drafting":
        return select_scenario_for_drafting(project_id, scenario_id)

    with get_cursor() as cur:
        cur.execute("SELECT scenario_purpose FROM er_scenarios WHERE id = %s AND project_id = %s", (scenario_id, project_id))
        row = cur.fetchone()
        if not row:
            return {"error": f"Scenario {scenario_id} not found for project {project_id}"}

        old_purpose = row["scenario_purpose"]
        if old_purpose == "selected_for_drafting":
            cur.execute("UPDATE user_projects SET selected_scenario_id = NULL WHERE id = %s", (project_id,))

        cur.execute("""
            UPDATE er_scenarios SET scenario_purpose = %s WHERE id = %s AND project_id = %s RETURNING id, name
        """, (purpose, scenario_id, project_id))
        updated = cur.fetchone()
        return {"scenario_id": updated["id"], "scenario_name": updated["name"], "purpose": purpose}


def get_selected_scenario(project_id):
    with get_cursor() as cur:
        cur.execute("""
            SELECT es.* FROM er_scenarios es
            JOIN user_projects up ON up.selected_scenario_id = es.id
            WHERE up.id = %s
        """, (project_id,))
        scenario = cur.fetchone()
        if not scenario:
            return None

        cur.execute("""
            SELECT * FROM er_scenario_years WHERE scenario_id = %s ORDER BY year_number
        """, (scenario["id"],))
        years = cur.fetchall()
        return {"scenario": scenario, "years": years}


def compare_scenarios(project_id, scenario_ids=None):
    with get_cursor() as cur:
        if scenario_ids:
            placeholders = ",".join(["%s"] * len(scenario_ids))
            cur.execute(f"""
                SELECT * FROM er_scenarios
                WHERE project_id = %s AND id IN ({placeholders})
                ORDER BY created_at DESC
            """, (project_id, *scenario_ids))
        else:
            cur.execute("""
                SELECT * FROM er_scenarios
                WHERE project_id = %s AND scenario_purpose != 'archived'
                ORDER BY created_at DESC
            """, (project_id,))
        scenarios = cur.fetchall()

        result = []
        for s in scenarios:
            summary = s.get("results_summary") or {}
            if isinstance(summary, str):
                try:
                    summary = json.loads(summary)
                except Exception:
                    summary = {}

            overrides = s.get("parameter_overrides") or {}
            if isinstance(overrides, str):
                try:
                    overrides = json.loads(overrides)
                except Exception:
                    overrides = {}
            clean_overrides = {k: v for k, v in overrides.items() if not str(k).startswith("__")}

            result.append({
                "id": s["id"],
                "name": s["name"],
                "scenario_purpose": s.get("scenario_purpose", "exploratory"),
                "is_baseline": s.get("is_baseline", False),
                "total_er": summary.get("total_er", 0),
                "average_annual_er": summary.get("average_annual_er", 0),
                "crediting_years": s.get("crediting_years", 7),
                "carbon_price_usd": s.get("carbon_price_usd"),
                "parameter_overrides": clean_overrides,
                "created_at": str(s.get("created_at", "")),
            })
        return {"scenarios": result, "count": len(result)}


def migrate_baseline_to_selected(project_id):
    with get_cursor() as cur:
        cur.execute("""
            SELECT selected_scenario_id FROM user_projects WHERE id = %s
        """, (project_id,))
        proj = cur.fetchone()
        if proj and proj.get("selected_scenario_id"):
            return {"migrated": False, "reason": "already_has_selected"}

        cur.execute("""
            SELECT id FROM er_scenarios
            WHERE project_id = %s AND scenario_purpose = 'selected_for_drafting'
            LIMIT 1
        """, (project_id,))
        if cur.fetchone():
            return {"migrated": False, "reason": "already_has_selected_purpose"}

        cur.execute("""
            SELECT id, name FROM er_scenarios
            WHERE project_id = %s AND is_baseline = true
              AND (scenario_purpose IS NULL OR scenario_purpose NOT IN ('selected_for_drafting'))
            ORDER BY calculated_at DESC
            LIMIT 1
        """, (project_id,))
        row = cur.fetchone()
        if row:
            scenario_id = row["id"]
            cur.execute("""
                UPDATE er_scenarios SET scenario_purpose = 'selected_for_drafting'
                WHERE id = %s
            """, (scenario_id,))
            cur.execute("""
                UPDATE user_projects SET selected_scenario_id = %s
                WHERE id = %s
            """, (scenario_id, project_id))
            return {"migrated": True, "scenario_id": scenario_id, "name": row["name"]}
    return {"migrated": False}


def run_sensitivity(project_id, param_key, variation_pct=20, steps=5):
    params = get_parameters_as_dict(project_id)
    if param_key not in params or params[param_key]["value"] is None:
        return {"error": f"Parameter {param_key} not found or has no value"}

    base_value = float(params[param_key]["value"])
    if base_value == 0:
        return {"error": f"Cannot run sensitivity on zero-value parameter {param_key}"}

    results = []
    for i in range(-steps, steps + 1):
        pct_change = (i / steps) * variation_pct
        test_value = base_value * (1 + pct_change / 100)
        overrides = {param_key: test_value}
        calc = run_scenario(project_id, parameter_overrides=overrides)
        if "error" not in calc:
            results.append({
                "pct_change": round(pct_change, 1),
                "param_value": round(test_value, 6),
                "total_er": calc["summary"]["total_er"],
                "annual_er": calc["summary"]["average_annual_er"],
            })

    base_calc = run_scenario(project_id)
    base_er = base_calc["summary"]["total_er"] if "error" not in base_calc else 0

    return {
        "param_key": param_key,
        "base_value": base_value,
        "base_er": base_er,
        "variation_pct": variation_pct,
        "results": results,
    }


def _add_finance(result, carbon_price, price_escalation=0, developer_share=100, buffer_pool=0, admin_fee=0):
    total_gross = 0
    total_deductions = 0
    total_net = 0

    for yr in result["years"]:
        y = yr["year_number"] - 1
        price = carbon_price * (1 + price_escalation / 100) ** y
        gross = yr["net_er"] * price
        buffer = gross * buffer_pool / 100
        admin = gross * admin_fee / 100
        deduct = buffer + admin
        net = (gross - deduct) * developer_share / 100

        yr["carbon_price"] = round(price, 2)
        yr["gross_revenue"] = round(gross, 2)
        yr["deductions"] = round(deduct, 2)
        yr["net_revenue"] = round(net, 2)
        yr["buffer_contribution"] = round(buffer, 2)

        total_gross += gross
        total_deductions += deduct
        total_net += net

    result["summary"]["total_gross_revenue"] = round(total_gross, 2)
    result["summary"]["total_deductions"] = round(total_deductions, 2)
    result["summary"]["total_net_revenue"] = round(total_net, 2)
    result["summary"]["carbon_price"] = carbon_price
    result["summary"]["developer_share_pct"] = developer_share


def export_er_to_excel(result, project_name="Project"):
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=12)
    param_header_font = Font(bold=True, size=10, color="FFFFFF")
    formula_font = Font(italic=True, size=9, color="666666")
    number_format_2dp = '0.00'
    number_format_6dp = '0.000000'
    teal_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    light_fill = PatternFill(start_color="F0FDFA", end_color="F0FDFA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws_params = wb.active
    ws_params.title = "Parameters"
    ws_params.append(["CarbonGPT - ER Calculation Workbook"])
    ws_params["A1"].font = Font(bold=True, size=14)
    ws_params.append([f"Project: {project_name}"])
    ws_params.append([f"Methodology: {result.get('summary', {}).get('methodology', 'N/A')}"])
    ws_params.append([])

    ws_params.append(["Parameter", "Value", "Unit", "Description"])
    for col in range(1, 5):
        cell = ws_params.cell(row=5, column=col)
        cell.font = param_header_font
        cell.fill = teal_fill
        cell.border = thin_border

    params_used = result.get("parameters_used", {})
    row = 6
    for key, info in params_used.items():
        if isinstance(info, dict):
            val = info.get("value", "")
            unit = info.get("unit", "")
            desc = info.get("description", "")
        else:
            val = info
            unit = ""
            desc = ""
        ws_params.append([key, val, unit, desc])
        for col in range(1, 5):
            ws_params.cell(row=row, column=col).border = thin_border
        if row % 2 == 0:
            for col in range(1, 5):
                ws_params.cell(row=row, column=col).fill = light_fill
        row += 1

    ws_params.column_dimensions['A'].width = 30
    ws_params.column_dimensions['B'].width = 18
    ws_params.column_dimensions['C'].width = 20
    ws_params.column_dimensions['D'].width = 45

    ws_steps = wb.create_sheet("Calculation Steps")
    ws_steps.append(["Step-by-Step Calculation Breakdown"])
    ws_steps["A1"].font = Font(bold=True, size=14)
    ws_steps.append([])

    ws_steps.append(["Step", "Description", "Formula / Calculation", "Result", "Unit"])
    for col in range(1, 6):
        cell = ws_steps.cell(row=3, column=col)
        cell.font = param_header_font
        cell.fill = teal_fill
        cell.border = thin_border

    steps = result.get("calculation_steps", [])
    row = 4
    for s in steps:
        step_num = s.get("step", "")
        name = s.get("name", "")
        formula = s.get("formula", "")
        val = s.get("value", s.get("value_baseline", ""))
        unit = s.get("unit", "")
        ws_steps.append([step_num, name, formula, val, unit])
        for col in range(1, 6):
            ws_steps.cell(row=row, column=col).border = thin_border
        ws_steps.cell(row=row, column=3).font = formula_font
        if isinstance(val, float):
            ws_steps.cell(row=row, column=4).number_format = number_format_6dp
        row += 1

    ws_steps.column_dimensions['A'].width = 8
    ws_steps.column_dimensions['B'].width = 45
    ws_steps.column_dimensions['C'].width = 70
    ws_steps.column_dimensions['D'].width = 18
    ws_steps.column_dimensions['E'].width = 20

    ws_years = wb.create_sheet("Year-by-Year Results")
    ws_years.append(["Year-by-Year Emission Reduction Calculations"])
    ws_years["A1"].font = Font(bold=True, size=14)
    ws_years.append([])

    year_headers = [
        "Year", "Calendar Year", "Usage Rate", "Active HH",
        "BE/hh (tCO2e)", "PE/hh (tCO2e)",
        "Baseline Emissions (tCO2e)", "BE Formula",
        "Project Emissions (tCO2e)", "PE Formula",
        "Gross ER (tCO2e)", "Leakage (tCO2e)", "Leakage Formula",
        "Net ER (tCO2e)", "Net ER Formula",
    ]
    ws_years.append(year_headers)
    for col in range(1, len(year_headers) + 1):
        cell = ws_years.cell(row=3, column=col)
        cell.font = param_header_font
        cell.fill = teal_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)

    years = result.get("years", [])
    row = 4
    for y in years:
        ws_years.append([
            y.get("year_number"),
            y.get("calendar_year"),
            y.get("usage_rate"),
            y.get("active_households"),
            y.get("be_per_hh"),
            y.get("pe_per_hh"),
            y.get("baseline_emissions"),
            y.get("baseline_formula", ""),
            y.get("project_emissions"),
            y.get("project_formula", ""),
            y.get("gross_er"),
            y.get("leakage"),
            y.get("leakage_formula", ""),
            y.get("net_er"),
            y.get("net_er_formula", ""),
        ])
        for col in range(1, len(year_headers) + 1):
            cell = ws_years.cell(row=row, column=col)
            cell.border = thin_border
            if isinstance(cell.value, float):
                cell.number_format = number_format_2dp
        for formula_col in [8, 10, 13, 15]:
            ws_years.cell(row=row, column=formula_col).font = formula_font
        if row % 2 == 0:
            for col in range(1, len(year_headers) + 1):
                ws_years.cell(row=row, column=col).fill = light_fill
        row += 1

    summary = result.get("summary", {})
    row += 1
    ws_years.cell(row=row, column=1, value="TOTALS").font = Font(bold=True, size=11)
    ws_years.cell(row=row, column=7, value=summary.get("total_baseline", 0)).font = Font(bold=True)
    ws_years.cell(row=row, column=7).number_format = number_format_2dp
    ws_years.cell(row=row, column=9, value=summary.get("total_project", 0)).font = Font(bold=True)
    ws_years.cell(row=row, column=9).number_format = number_format_2dp
    ws_years.cell(row=row, column=12, value=summary.get("total_leakage", 0)).font = Font(bold=True)
    ws_years.cell(row=row, column=12).number_format = number_format_2dp
    ws_years.cell(row=row, column=14, value=summary.get("total_er", 0)).font = Font(bold=True)
    ws_years.cell(row=row, column=14).number_format = number_format_2dp
    for col in range(1, len(year_headers) + 1):
        ws_years.cell(row=row, column=col).border = Border(top=Side(style='double'))

    for col_letter in ['A','B','C','D','E','F']:
        ws_years.column_dimensions[col_letter].width = 14
    for col_letter in ['G','H','I','J','K','L','M','N','O']:
        ws_years.column_dimensions[col_letter].width = 22

    deployment_timeline = result.get("deployment_timeline")
    if deployment_timeline:
        ws_deploy = wb.create_sheet("Deployment & Cohort")
        ws_deploy.append(["Deployment & Technology Dynamics"])
        ws_deploy["A1"].font = Font(bold=True, size=14)
        ws_deploy.append([])

        deploy_headers = [
            "Year", "Calendar Year", "Deployed", "Cumulative Deployed",
            "Active Units", "Surviving Units", "Effectively Used",
            "Net ER (tCO2e)", "Cumulative ER (tCO2e)",
        ]
        ws_deploy.append(deploy_headers)
        for col in range(1, len(deploy_headers) + 1):
            cell = ws_deploy.cell(row=3, column=col)
            cell.font = param_header_font
            cell.fill = teal_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

        drow = 4
        for dt in deployment_timeline:
            ws_deploy.append([
                dt.get("year_number", ""),
                dt.get("year", ""),
                dt.get("deployed", 0),
                dt.get("cumulative_deployed", 0),
                dt.get("active", 0),
                dt.get("surviving", 0),
                dt.get("effectively_used", 0),
                dt.get("net_er", 0),
                dt.get("cumulative_er", 0),
            ])
            for col in range(1, len(deploy_headers) + 1):
                cell = ws_deploy.cell(row=drow, column=col)
                cell.border = thin_border
                if isinstance(cell.value, float):
                    cell.number_format = number_format_2dp
            if drow % 2 == 0:
                for col in range(1, len(deploy_headers) + 1):
                    ws_deploy.cell(row=drow, column=col).fill = light_fill
            drow += 1

        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws_deploy.column_dimensions[col_letter].width = 18

    if "total_gross_revenue" in summary:
        ws_fin = wb.create_sheet("Carbon Finance")
        ws_fin.append(["Carbon Finance Projections"])
        ws_fin["A1"].font = Font(bold=True, size=14)
        ws_fin.append([])
        ws_fin.append(["Carbon Price ($/tCO2e)", summary.get("carbon_price", 0)])
        ws_fin.append(["Developer Share (%)", summary.get("developer_share_pct", 100)])
        ws_fin.append(["Total Gross Revenue ($)", summary.get("total_gross_revenue", 0)])
        ws_fin.append(["Total Deductions ($)", summary.get("total_deductions", 0)])
        ws_fin.append(["Total Net Revenue ($)", summary.get("total_net_revenue", 0)])
        ws_fin.column_dimensions['A'].width = 30
        ws_fin.column_dimensions['B'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _pval(params, key, default=0.0):
    if key in params:
        v = params[key]
        if isinstance(v, dict):
            v = v.get("value")
        if v is not None:
            try:
                return float(v)
            except (ValueError, TypeError):
                pass
    return default


def _ptext(params, key, default=""):
    if key in params:
        v = params[key]
        if isinstance(v, dict):
            v = v.get("value")
        if v is not None:
            return str(v)
    return default
